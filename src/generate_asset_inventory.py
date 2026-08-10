#!/usr/bin/env python3
"""
Asset Inventory Generator for Canadian Families in Ontario
Generates Markdown, Excel, and HTML Dashboard outputs.
"""

import json
import argparse
import base64
import hashlib
import hmac
import re
from datetime import datetime, date
from pathlib import Path
from typing import Optional

from translations import (
    _,
    CATEGORY_TRANSLATIONS,
    UI_TRANSLATIONS,
    translate_assets,
    translate_field_definitions,
)


def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b"=").decode("ascii")


def sign_license(payload: dict, secret: str) -> str:
    """Produce a signed license: base64url(json).hmac-sha256(base64url(json), secret).

    The HMAC is computed over the base64url body string so the browser can verify
    it without decoding (avoids padding/url-safe mismatches).
    """
    body = json.dumps(payload, separators=(",", ":")).encode("utf-8")
    body_b64 = _b64url(body)
    sig = hmac.new(secret.encode("utf-8"), body_b64.encode("ascii"), hashlib.sha256).digest()
    return body_b64 + "." + _b64url(sig)


# =============================================================================
# ASSET CATEGORIES AND DEFINITIONS
# =============================================================================

ASSET_CATEGORIES = {
    "Cash & Cash Equivalents": {
        "icon": "💰",
        "color": "#00b894",
        "items": [
            {"name": "Chequing Account", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO", "Scotiabank", "CIBC"]},
            {"name": "High-Interest Savings Account", "institution_type": "Bank", "typical_institutions": ["EQ Bank", "Tangerine", "Simplii", "Motus Bank"]},
            {"name": "Regular Savings Account", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO", "Scotiabank", "CIBC"]},
            {"name": "Money Market Account", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO"]},
            {"name": "Physical Cash", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Foreign Currency Holdings", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "Knightsbridge FX"]},
            {"name": "Money Order", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Certified Cheque", "institution_type": "Bank", "typical_institutions": []},
            {"name": "Prepaid Debit/Credit Card", "institution_type": "Financial", "typical_institutions": ["Koho", "Stack", "Mogo"]},
            {"name": "Cashable GIC", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO", "EQ Bank"]},
        ]
    },
    "Fixed Income Investments": {
        "icon": "📈",
        "color": "#0984e3",
        "items": [
            {"name": "Canada Savings Bonds (CSB)", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "Canada Premium Bonds (CPB)", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "Provincial Savings Bonds", "institution_type": "Government", "typical_institutions": ["Province of Ontario"]},
            {"name": "Government of Canada Bonds", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "Provincial Government Bonds", "institution_type": "Government", "typical_institutions": ["Province of Ontario"]},
            {"name": "Municipal Bonds", "institution_type": "Government", "typical_institutions": ["City of Toronto", "Province of Ontario"]},
            {"name": "Corporate Bonds", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR", "Wealthsimple"]},
            {"name": "Guaranteed Investment Certificates (GICs)", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO", "EQ Bank", "Oaken Financial"]},
            {"name": "Strip Bonds", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"]},
            {"name": "Strip Coupons", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"]},
            {"name": "Mortgage-Backed Securities (MBS)", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"]},
            {"name": "Treasury Bills (T-Bills)", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "High-Yield Bonds", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"]},
            {"name": "Convertible Bonds", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"]},
            {"name": "Foreign Bonds", "institution_type": "Brokerage", "typical_institutions": ["IBKR", "Questrade"]},
        ]
    },
    "Equities & Investment Funds": {
        "icon": "📊",
        "color": "#6c5ce7",
        "items": [
            {"name": "Individual Canadian Stocks (TSX)", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR", "Wealthsimple", "TD Direct Investing"]},
            {"name": "Individual US Stocks (NYSE/NASDAQ)", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR", "Wealthsimple"]},
            {"name": "Individual International Stocks", "institution_type": "Brokerage", "typical_institutions": ["IBKR", "Questrade"]},
            {"name": "Canadian Equity ETFs", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "IBKR"]},
            {"name": "US Equity ETFs", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "IBKR"]},
            {"name": "International Equity ETFs", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "IBKR"]},
            {"name": "Sector ETFs", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"]},
            {"name": "Bond ETFs", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "IBKR"]},
            {"name": "Canadian Mutual Funds", "institution_type": "Fund Company", "typical_institutions": ["Fidelity", "RBC Global Asset Management", "TD Asset Management", "BMO Asset Management"]},
            {"name": "US Mutual Funds", "institution_type": "Fund Company", "typical_institutions": ["Vanguard", "Fidelity", "T. Rowe Price"]},
            {"name": "Index Funds", "institution_type": "Fund Company", "typical_institutions": ["Vanguard", "iShares", "BMO"]},
            {"name": "Hedge Funds", "institution_type": "Fund Company", "typical_institutions": ["Various"]},
            {"name": "Segregated Funds", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life", "Industrial Alliance"]},
            {"name": "Principal-Protected Notes (PPNs)", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO", "Scotiabank"]},
            {"name": "Structured Notes", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO"]},
        ]
    },
    "Registered Accounts — Canada": {
        "icon": "🏦",
        "color": "#e17055",
        "items": [
            {"name": "TFSA (Tax-Free Savings Account)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "TD", "RBC"], "tax_treatment": "Tax-free"},
            {"name": "RRSP (Registered Retirement Savings Plan)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "TD", "RBC"], "tax_treatment": "Tax-deferred"},
            {"name": "RESP (Registered Education Savings Plan)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "TD", "RBC"], "tax_treatment": "Tax-deferred"},
            {"name": "RRIF (Registered Retirement Income Fund)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "TD", "RBC"], "tax_treatment": "Taxable on withdrawal"},
            {"name": "LIRA (Locked-In Retirement Account)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "TD", "RBC"], "tax_treatment": "Tax-deferred"},
            {"name": "LIF (Locked-In Income Fund)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "TD", "RBC"], "tax_treatment": "Taxable on withdrawal"},
            {"name": "LRIF (Locked-In Retirement Income Fund)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "TD", "RBC"], "tax_treatment": "Taxable on withdrawal"},
            {"name": "PRIF (Prescribed Retirement Income Fund)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "TD"], "tax_treatment": "Taxable on withdrawal"},
            {"name": "RPP (Registered Pension Plan)", "institution_type": "Employer", "typical_institutions": ["Employer plan"], "tax_treatment": "Tax-deferred"},
            {"name": "DPSP (Deferred Profit Sharing Plan)", "institution_type": "Employer", "typical_institutions": ["Employer plan"], "tax_treatment": "Tax-deferred"},
            {"name": "RDSP (Registered Disability Savings Plan)", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO"], "tax_treatment": "Tax-deferred"},
            {"name": "FHSA (First Home Savings Account)", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "TD", "RBC"], "tax_treatment": "Tax-free"},
            {"name": "Non-Registered Investment Account", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "IBKR"], "tax_treatment": "Taxable"},
            {"name": "Margin Account", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"], "tax_treatment": "Taxable"},
            {"name": "Cash Brokerage Account", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple"], "tax_treatment": "Taxable"},
        ]
    },
    "Employment & Compensation Assets": {
        "icon": "💼",
        "color": "#00cec9",
        "items": [
            {"name": "Stock Options (Vested)", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "Stock Options (Unvested)", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "RSUs — Restricted Stock Units (Vested)", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "RSUs — Restricted Stock Units (Unvested)", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "DSUs — Deferred Stock Units", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "ESPP — Employee Stock Purchase Plan", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "Performance Shares/Bonuses", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "Deferred Compensation", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "Severance Entitlements", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "Vacation Pay Owing", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "Commissions Owing", "institution_type": "Employer", "typical_institutions": ["Employer"]},
        ]
    },
    "Pension Benefits": {
        "icon": "🏛️",
        "color": "#636e72",
        "items": [
            {"name": "CPP — Canada Pension Plan Credits", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "QPP — Quebec Pension Plan Credits", "institution_type": "Government", "typical_institutions": ["Government of Quebec"]},
            {"name": "OAS — Old Age Security Credits", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "GIS — Guaranteed Income Supplement", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "Employer Defined-Benefit Pension", "institution_type": "Employer", "typical_institutions": ["Employer plan"]},
            {"name": "Employer Defined-Contribution Pension", "institution_type": "Employer", "typical_institutions": ["Employer plan"]},
            {"name": "Group RRSP", "institution_type": "Employer", "typical_institutions": ["Employer plan"]},
            {"name": "Group TFSA", "institution_type": "Employer", "typical_institutions": ["Employer plan"]},
            {"name": "Public Sector Pension", "institution_type": "Government", "typical_institutions": ["Federal/Provincial/Municipal"]},
            {"name": "Federal Government Pension", "institution_type": "Government", "typical_institutions": ["Government of Canada"]},
            {"name": "Provincial Government Pension", "institution_type": "Government", "typical_institutions": ["Province of Ontario"]},
            {"name": "Municipal Government Pension", "institution_type": "Government", "typical_institutions": ["City of Toronto"]},
            {"name": "Military Pension", "institution_type": "Government", "typical_institutions": ["Department of National Defence"]},
            {"name": "RCMP Pension", "institution_type": "Government", "typical_institutions": ["Royal Canadian Mounted Police"]},
            {"name": "Police/Fire Pension", "institution_type": "Government", "typical_institutions": ["Local police/fire services"]},
            {"name": "Teacher Pension", "institution_type": "Government", "typical_institutions": ["Ontario Teachers' Pension Plan"]},
            {"name": "Nurse Pension", "institution_type": "Government", "typical_institutions": ["HOOPP"]},
        ]
    },
    "Insurance Products": {
        "icon": "🛡️",
        "color": "#fdcb6e",
        "items": [
            {"name": "Term Life Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life", "Industrial Alliance"]},
            {"name": "Whole Life Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Universal Life Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Critical Illness Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Disability Insurance (Individual)", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Disability Insurance (Group/Employer)", "institution_type": "Insurance", "typical_institutions": ["Employer plan"]},
            {"name": "Health Insurance (Private Supplemental)", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Green Shield"]},
            {"name": "Dental Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Vision Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life"]},
            {"name": "Travel Medical Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Blue Cross"]},
            {"name": "Pet Insurance", "institution_type": "Insurance", "typical_institutions": ["Petsecure", "Trupanion", "Pets Plus Us"]},
            {"name": "Homeowners Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva", "Desjardins", "The Co-operators"]},
            {"name": "Condo Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva", "Desjardins"]},
            {"name": "Tenants Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva", "Desjardins"]},
            {"name": "Auto Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva", "Desjardins", "Belairdirect"]},
            {"name": "Boat/Watercraft Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva"]},
            {"name": "RV/Camper Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva"]},
            {"name": "Motorcycle Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva"]},
            {"name": "Snowmobile Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva"]},
            {"name": "ATV Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva"]},
            {"name": "Classic/Collector Car Insurance", "institution_type": "Insurance", "typical_institutions": ["Hagerty", "Intact"]},
            {"name": "Commercial Property Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva", "Zurich"]},
            {"name": "Business Liability Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva", "Zurich"]},
            {"name": "Key Person Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Buy-Sell Agreement Insurance", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Annuities (Fixed)", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life", "Canada Life"]},
            {"name": "Annuities (Variable)", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life"]},
            {"name": "Registered Annuities (RRIF-eligible)", "institution_type": "Insurance", "typical_institutions": ["Manulife", "Sun Life"]},
        ]
    },
    "Real Estate": {
        "icon": "🏠",
        "color": "#e17055",
        "items": [
            {"name": "Primary Residence", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Secondary/Vacation Property", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Rental Property (Residential)", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Rental Property (Commercial)", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Investment Property", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Agricultural Land", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Vacant Land", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Timeshare", "institution_type": "N/A", "typical_institutions": ["Marriott Vacation Club", "Hilton Grand Vacations"]},
            {"name": "Co-Ownership Interest", "institution_type": "N/A", "typical_institutions": []},
            {"name": "REITs (Public)", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple"]},
            {"name": "REITs (Private)", "institution_type": "Fund Company", "typical_institutions": ["Artis REIT", "RioCan", "Canadian REIT"]},
            {"name": "Mortgage Investments (Private)", "institution_type": "Mortgage Investment Corp", "typical_institutions": ["Various MICs"]},
            {"name": "Real Estate Partnerships", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Farmland", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Waterfront Property", "institution_type": "N/A", "typical_institutions": []},
        ]
    },
    "Vehicles & Transportation": {
        "icon": "🚗",
        "color": "#74b9ff",
        "items": [
            {"name": "Car/Truck", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Motorcycle", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Boat/Yacht", "institution_type": "N/A", "typical_institutions": []},
            {"name": "RV/Camper/Trailer", "institution_type": "N/A", "typical_institutions": []},
            {"name": "ATV/UTV", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Snowmobile", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Personal Watercraft (Jet Ski, Sea-Doo)", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Aircraft", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Classic/Collector Vehicle", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Electric Vehicle", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Used Vehicle", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Commercial Vehicle", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Recreational Trailer", "institution_type": "N/A", "typical_institutions": []},
        ]
    },
    "Personal Property — Valuables": {
        "icon": "💎",
        "color": "#fd79a8",
        "items": [
            {"name": "Jewelry", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Watches", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Artwork (Paintings, Sculptures)", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Antiques", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Coin Collections", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Stamp Collections", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Wine/Spirits Collection", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Rare Books", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Sports Memorabilia", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Furs", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Precious Metals (Gold, Silver, Platinum)", "institution_type": "Dealer", "typical_institutions": ["Kitco", "Silver Gold Bull"]},
            {"name": "Gems/Diamonds", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Designer Handbags/Accessories", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Musical Instruments", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Firearms (Licensed)", "institution_type": "N/A", "typical_institutions": []},
        ]
    },
    "Household & Electronics": {
        "icon": "🖥️",
        "color": "#a29bfe",
        "items": [
            {"name": "Furniture", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Major Appliances", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Computers/Laptops", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Smartphones/Tablets", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Gaming Consoles/Systems", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Audio/Video Equipment", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Home Theater Systems", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Smart Home Devices", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Kitchen Equipment", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Power Tools", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Garden/Yard Equipment", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Exercise/Fitness Equipment", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Camping/Outdoor Gear", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Sports Equipment", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Photography/Videography Equipment", "institution_type": "N/A", "typical_institutions": []},
        ]
    },
    "Cryptocurrencies": {
        "icon": "₿",
        "color": "#f39c12",
        "items": [
            {"name": "Bitcoin (BTC)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Shakepay", "Coinbase", "Kraken"]},
            {"name": "Ethereum (ETH)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Shakepay", "Coinbase", "Kraken"]},
            {"name": "Litecoin (LTC)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Shakepay", "Coinbase"]},
            {"name": "Ripple/XRP (XRP)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase", "Kraken"]},
            {"name": "Bitcoin Cash (BCH)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Cardano (ADA)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase", "Kraken"]},
            {"name": "Polkadot (DOT)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase", "Kraken"]},
            {"name": "Solana (SOL)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase", "Kraken"]},
            {"name": "Dogecoin (DOGE)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Shakepay", "Coinbase"]},
            {"name": "Chainlink (LINK)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase", "Kraken"]},
            {"name": "Stellar (XLM)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Monero (XMR)", "institution_type": "Exchange", "typical_institutions": ["Kraken"]},
            {"name": "EOS", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Tron (TRX)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Avalanche (AVAX)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase", "Kraken"]},
            {"name": "Polygon (MATIC)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Uniswap (UNI)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Aave (AAVE)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Compound (COMP)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Maker (MKR)", "institution_type": "Exchange", "typical_institutions": ["Coinbase", "Kraken"]},
            {"name": "Yearn.finance (YFI)", "institution_type": "Exchange", "typical_institutions": ["Coinbase", "Kraken"]},
            {"name": "SushiSwap (SUSHI)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Synthetix (SNX)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Curve (CRV)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "1inch (1INCH)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "PancakeSwap (CAKE)", "institution_type": "Exchange", "typical_institutions": ["Binance"]},
            {"name": "Terra/Luna (LUNA)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Algorand (ALGO)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Cosmos (ATOM)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Near Protocol (NEAR)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase"]},
            {"name": "Fantom (FTM)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Internet Computer (ICP)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Hedera (HBAR)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "VeChain (VET)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Filecoin (FIL)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Theta (THETA)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "The Graph (GRT)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Axie Infinity (AXS)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Decentraland (MANA)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Sandbox (SAND)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Enjin Coin (ENJ)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
        ]
    },
    "Stablecoins": {
        "icon": "💲",
        "color": "#00b894",
        "items": [
            {"name": "USDC (USD Coin)", "institution_type": "Exchange", "typical_institutions": ["Newton", "Coinbase", "Kraken"]},
            {"name": "USDT (Tether)", "institution_type": "Exchange", "typical_institutions": ["Coinbase", "Kraken", "Binance"]},
            {"name": "DAI", "institution_type": "DeFi", "typical_institutions": ["Uniswap", "Compound"]},
            {"name": "BUSD (Binance USD)", "institution_type": "Exchange", "typical_institutions": ["Binance"]},
            {"name": "TrueUSD (TUSD)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Pax Dollar (USDP)", "institution_type": "Exchange", "typical_institutions": ["Coinbase"]},
            {"name": "Frax (FRAX)", "institution_type": "DeFi", "typical_institutions": ["Uniswap"]},
            {"name": "Liquity USD (LUSD)", "institution_type": "DeFi", "typical_institutions": ["Liquity"]},
            {"name": "Gemini Dollar (GUSD)", "institution_type": "Exchange", "typical_institutions": ["Gemini"]},
            {"name": "sUSD", "institution_type": "DeFi", "typical_institutions": ["Synthetix"]},
        ]
    },
    "DeFi & Staking": {
        "icon": "🔗",
        "color": "#6c5ce7",
        "items": [
            {"name": "Staked ETH", "institution_type": "DeFi", "typical_institutions": ["Lido", "Coinbase", "Rocket Pool"]},
            {"name": "Staked SOL", "institution_type": "DeFi", "typical_institutions": ["Marinade", "Jito"]},
            {"name": "Staked ADA", "institution_type": "DeFi", "typical_institutions": ["Various pools"]},
            {"name": "Staked DOT", "institution_type": "DeFi", "typical_institutions": ["Various validators"]},
            {"name": "Liquidity Pool Tokens", "institution_type": "DeFi", "typical_institutions": ["Uniswap", "SushiSwap", "PancakeSwap"]},
            {"name": "Yield Farming Positions", "institution_type": "DeFi", "typical_institutions": ["Various protocols"]},
            {"name": "Lending Positions (Aave, Compound)", "institution_type": "DeFi", "typical_institutions": ["Aave", "Compound"]},
            {"name": "Liquity Positions", "institution_type": "DeFi", "typical_institutions": ["Liquity"]},
            {"name": "LP Tokens (Uniswap)", "institution_type": "DeFi", "typical_institutions": ["Uniswap"]},
            {"name": "LP Tokens (SushiSwap)", "institution_type": "DeFi", "typical_institutions": ["SushiSwap"]},
            {"name": "LP Tokens (PancakeSwap)", "institution_type": "DeFi", "typical_institutions": ["PancakeSwap"]},
            {"name": "Vault Positions", "institution_type": "DeFi", "typical_institutions": ["Yearn", "Convex"]},
            {"name": "Governance Tokens", "institution_type": "DeFi", "typical_institutions": ["Various DAOs"]},
            {"name": "Wrapper Tokens (WBTC, WETH)", "institution_type": "DeFi", "typical_institutions": ["Various bridges"]},
            {"name": "Rebasing Tokens", "institution_type": "DeFi", "typical_institutions": ["Ampleforth", "Empty Set Dollar"]},
            {"name": "Flash Loan Positions", "institution_type": "DeFi", "typical_institutions": ["Aave"]},
        ]
    },
    "NFTs (Non-Fungible Tokens)": {
        "icon": "🖼️",
        "color": "#fd79a8",
        "items": [
            {"name": "Digital Art NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea", "Foundation", "SuperRare"]},
            {"name": "PFP (Profile Picture) NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea"]},
            {"name": "Gaming NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea", "Immutable X"]},
            {"name": "Metaverse Land/Property NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea", "Decentraland", "Sandbox"]},
            {"name": "Music NFTs", "institution_type": "Marketplace", "typical_institutions": ["Sound.xyz", "OpenSea"]},
            {"name": "Video NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea"]},
            {"name": "Domain Name NFTs (.eth, .crypto)", "institution_type": "Marketplace", "typical_institutions": ["ENS", "Unstoppable Domains"]},
            {"name": "Membership NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea"]},
            {"name": "Sports Collectible NFTs", "institution_type": "Marketplace", "typical_institutions": ["NBA Top Shot", "Sorare"]},
            {"name": "Trading Card NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea"]},
            {"name": "Generative Art NFTs", "institution_type": "Marketplace", "typical_institutions": ["Art Blocks", "fxhash"]},
            {"name": "Photography NFTs", "institution_type": "Marketplace", "typical_institutions": ["OpenSea", "Foundation"]},
        ]
    },
    "Crypto Wallets & Infrastructure": {
        "icon": "🔐",
        "color": "#636e72",
        "items": [
            {"name": "Hardware Wallet (Ledger, Trezor, etc.)", "institution_type": "Hardware", "typical_institutions": ["Ledger", "Trezor", "Keystone"]},
            {"name": "Software Wallet (MetaMask, Trust Wallet, etc.)", "institution_type": "Software", "typical_institutions": ["MetaMask", "Trust Wallet", "Phantom"]},
            {"name": "Exchange Account Balances", "institution_type": "Exchange", "typical_institutions": ["Newton", "Shakepay", "Coinbase", "Kraken"]},
            {"name": "Mining Equipment (GPU Miners, ASIC Miners)", "institution_type": "Hardware", "typical_institutions": ["Bitmain", "MicroBT"]},
            {"name": "Mining Pool Shares", "institution_type": "Mining", "typical_institutions": ["Various pools"]},
            {"name": "Validator Node Deposits", "institution_type": "DeFi", "typical_institutions": ["Various chains"]},
            {"name": "Lightning Network Channels", "institution_type": "DeFi", "typical_institutions": ["Various nodes"]},
            {"name": "Cross-Chain Bridge Positions", "institution_type": "DeFi", "typical_institutions": ["Various bridges"]},
        ]
    },
    "Digital Assets — Online Presence": {
        "icon": "🌐",
        "color": "#0984e3",
        "items": [
            {"name": "Domain Names", "institution_type": "Registrar", "typical_institutions": ["GoDaddy", "Namecheap", "Cloudflare"]},
            {"name": "Website (Personal Blog)", "institution_type": "Hosting", "typical_institutions": ["WordPress", "Squarespace", "Wix"]},
            {"name": "Website (Monetized Content)", "institution_type": "Hosting", "typical_institutions": ["WordPress", "Ghost"]},
            {"name": "Website (E-Commerce)", "institution_type": "Hosting", "typical_institutions": ["Shopify", "WooCommerce"]},
            {"name": "YouTube Channel", "institution_type": "Platform", "typical_institutions": ["YouTube"]},
            {"name": "YouTube Channel (Monetized)", "institution_type": "Platform", "typical_institutions": ["YouTube"]},
            {"name": "TikTok Account", "institution_type": "Platform", "typical_institutions": ["TikTok"]},
            {"name": "Instagram Account", "institution_type": "Platform", "typical_institutions": ["Instagram"]},
            {"name": "Twitter/X Account", "institution_type": "Platform", "typical_institutions": ["X/Twitter"]},
            {"name": "Twitch Channel", "institution_type": "Platform", "typical_institutions": ["Twitch"]},
            {"name": "Podcast", "institution_type": "Platform", "typical_institutions": ["Spotify", "Apple Podcasts"]},
            {"name": "Newsletter Subscriber List", "institution_type": "Platform", "typical_institutions": ["Substack", "ConvertKit", "Mailchimp"]},
            {"name": "Email List", "institution_type": "Platform", "typical_institutions": ["Various ESPs"]},
            {"name": "Online Community/Forum", "institution_type": "Platform", "typical_institutions": ["Discourse", "Circle"]},
            {"name": "Discord Server", "institution_type": "Platform", "typical_institutions": ["Discord"]},
            {"name": "Telegram Group", "institution_type": "Platform", "typical_institutions": ["Telegram"]},
            {"name": "Substack/Paid Newsletter", "institution_type": "Platform", "typical_institutions": ["Substack"]},
        ]
    },
    "Digital Assets — Online Businesses & Income": {
        "icon": "🛒",
        "color": "#00cec9",
        "items": [
            {"name": "Shopify Store", "institution_type": "E-Commerce", "typical_institutions": ["Shopify"]},
            {"name": "Amazon FBA Business", "institution_type": "E-Commerce", "typical_institutions": ["Amazon"]},
            {"name": "Etsy Shop", "institution_type": "E-Commerce", "typical_institutions": ["Etsy"]},
            {"name": "eBay Store", "institution_type": "E-Commerce", "typical_institutions": ["eBay"]},
            {"name": "Digital Product Store", "institution_type": "E-Commerce", "typical_institutions": ["Gumroad", "Lemonsqueezy"]},
            {"name": "Online Course (Hosted)", "institution_type": "EdTech", "typical_institutions": ["Teachable", "Kajabi", "Thinkific"]},
            {"name": "SaaS (Software as a Service)", "institution_type": "Tech", "typical_institutions": ["Various"]},
            {"name": "Mobile App", "institution_type": "Tech", "typical_institutions": ["App Store", "Google Play"]},
            {"name": "Browser Extension", "institution_type": "Tech", "typical_institutions": ["Chrome Web Store"]},
            {"name": "API Service", "institution_type": "Tech", "typical_institutions": ["Various"]},
            {"name": "Freelance Platform Accounts", "institution_type": "Platform", "typical_institutions": ["Upwork", "Fiverr", "Toptal"]},
            {"name": "Print-on-Demand Store", "institution_type": "E-Commerce", "typical_institutions": ["Redbubble", "Teespring", "Merch by Amazon"]},
            {"name": "Affiliate Marketing Sites", "institution_type": "Marketing", "typical_institutions": ["Various"]},
            {"name": "Dropshipping Business", "institution_type": "E-Commerce", "typical_institutions": ["Various"]},
        ]
    },
    "Digital Assets — Content & IP": {
        "icon": "📚",
        "color": "#e17055",
        "items": [
            {"name": "E-Books", "institution_type": "Publisher", "typical_institutions": ["Amazon KDP", "Gumroad"]},
            {"name": "Digital Music/Albums", "institution_type": "Distributor", "typical_institutions": ["DistroKid", "TuneCore"]},
            {"name": "Online Courses/Videos", "institution_type": "EdTech", "typical_institutions": ["Udemy", "Skillshare"]},
            {"name": "Digital Templates", "institution_type": "Marketplace", "typical_institutions": ["Creative Market", "Etsy"]},
            {"name": "Digital Tools/Software", "institution_type": "Tech", "typical_institutions": ["Various"]},
            {"name": "Stock Photos/Videos", "institution_type": "Marketplace", "typical_institutions": ["Shutterstock", "Adobe Stock"]},
            {"name": "Graphic Designs", "institution_type": "Marketplace", "typical_institutions": ["Creative Market", "Envato"]},
            {"name": "Fonts/Typefaces", "institution_type": "Marketplace", "typical_institutions": ["MyFonts", "Creative Market"]},
            {"name": "Digital Plugins", "institution_type": "Tech", "typical_institutions": ["Various"]},
            {"name": "WordPress Themes", "institution_type": "Marketplace", "typical_institutions": ["ThemeForest", "Elegant Themes"]},
            {"name": "3D Models", "institution_type": "Marketplace", "typical_institutions": ["TurboSquid", "CGTrader"]},
            {"name": "Game Assets", "institution_type": "Marketplace", "typical_institutions": ["Unity Asset Store", "Unreal Marketplace"]},
            {"name": "AI Models/Prompts", "institution_type": "Tech", "typical_institutions": ["PromptBase", "Various"]},
        ]
    },
    "Digital Assets — Accounts & Subscriptions": {
        "icon": "🔑",
        "color": "#636e72",
        "items": [
            {"name": "Cloud Storage (with value/content)", "institution_type": "Cloud", "typical_institutions": ["Google Drive", "iCloud", "Dropbox"]},
            {"name": "Premium Software Licenses", "institution_type": "Software", "typical_institutions": ["Adobe", "Microsoft", "Various"]},
            {"name": "App Store Accounts (Apple/Google)", "institution_type": "Platform", "typical_institutions": ["Apple", "Google"]},
            {"name": "Steam Accounts (with game library)", "institution_type": "Gaming", "typical_institutions": ["Steam"]},
            {"name": "Online Gaming Accounts", "institution_type": "Gaming", "typical_institutions": ["Various"]},
            {"name": "Virtual World Accounts", "institution_type": "Virtual", "typical_institutions": ["Second Life", "VRChat"]},
            {"name": "Domain Registrar Accounts", "institution_type": "Registrar", "typical_institutions": ["GoDaddy", "Namecheap"]},
            {"name": "Hosting Accounts", "institution_type": "Hosting", "typical_institutions": ["Bluehost", "SiteGround"]},
            {"name": "VPN Subscriptions (Prepaid)", "institution_type": "VPN", "typical_institutions": ["NordVPN", "ExpressVPN"]},
            {"name": "Streaming Service Gift Balances", "institution_type": "Streaming", "typical_institutions": ["Netflix", "Spotify", "Disney+"]},
            {"name": "Digital Wallet Balances", "institution_type": "Payment", "typical_institutions": ["PayPal", "Venmo", "Apple Pay"]},
            {"name": "In-Game Currency/Items", "institution_type": "Gaming", "typical_institutions": ["Fortnite", "Roblox", "Various"]},
            {"name": "Virtual Items/Cosmetics", "institution_type": "Gaming", "typical_institutions": ["Various"]},
        ]
    },
    "Intellectual Property": {
        "icon": "📝",
        "color": "#fdcb6e",
        "items": [
            {"name": "Patents", "institution_type": "Legal", "typical_institutions": ["CIPO (Canadian Intellectual Property Office)"]},
            {"name": "Patent Applications", "institution_type": "Legal", "typical_institutions": ["CIPO"]},
            {"name": "Trademarks", "institution_type": "Legal", "typical_institutions": ["CIPO"]},
            {"name": "Copyrights", "institution_type": "Legal", "typical_institutions": ["Canadian Copyright Office"]},
            {"name": "Trade Secrets", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Franchise Rights", "institution_type": "Legal", "typical_institutions": ["Franchisor"]},
            {"name": "License Agreements (Outgoing)", "institution_type": "Legal", "typical_institutions": ["Various"]},
            {"name": "License Agreements (Incoming)", "institution_type": "Legal", "typical_institutions": ["Various"]},
            {"name": "Royalty Agreements", "institution_type": "Legal", "typical_institutions": ["Various"]},
            {"name": "Music Publishing Rights", "institution_type": "Legal", "typical_institutions": ["SOCAN", "CMRRA"]},
            {"name": "Film/TV Rights", "institution_type": "Legal", "typical_institutions": ["Various"]},
            {"name": "Book Publishing Rights", "institution_type": "Legal", "typical_institutions": ["Various publishers"]},
            {"name": "Software Copyrights", "institution_type": "Legal", "typical_institutions": ["CIPO"]},
            {"name": "Database Rights", "institution_type": "Legal", "typical_institutions": ["Various"]},
            {"name": "Industrial Designs", "institution_type": "Legal", "typical_institutions": ["CIPO"]},
            {"name": "Plant Variety Rights", "institution_type": "Legal", "typical_institutions": ["CIPO"]},
        ]
    },
    "Business Assets": {
        "icon": "🏢",
        "color": "#00b894",
        "items": [
            {"name": "Sole Proprietorship", "institution_type": "N/A", "typical_institutions": []},
            {"name": "General Partnership Interest", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Limited Partnership Interest", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Corporation Shares (Private)", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Corporation Shares (Public)", "institution_type": "N/A", "typical_institutions": []},
            {"name": "LLC Membership", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Goodwill", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Client/Customer Lists", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Contracts/Agreements", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Equipment", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Inventory", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Accounts Receivable", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Bank Accounts", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO", "Scotiabank"]},
            {"name": "Business Investments", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Vehicles", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Real Estate", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Intellectual Property", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Business Licenses/Permits", "institution_type": "Government", "typical_institutions": ["Municipal/Provincial"]},
            {"name": "Business Insurance", "institution_type": "Insurance", "typical_institutions": ["Intact", "Aviva", "Zurich"]},
            {"name": "Business Security Deposits", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Vendor Relationships", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Distributor Agreements", "institution_type": "N/A", "typical_institutions": []},
        ]
    },
    "Government Benefits & Tax Credits": {
        "icon": "🏛️",
        "color": "#6c5ce7",
        "items": [
            {"name": "GST/HST Credit Entitlement", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Ontario Trillium Benefit", "institution_type": "Government", "typical_institutions": ["Ontario Ministry of Finance"]},
            {"name": "Canada Workers Benefit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Canada Child Benefit (CCB)", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Canada Training Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Canada Caregiver Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Disability Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Medical Expense Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Charitable Donation Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Teacher/EC Educator School Supply Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Home Accessibility Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Digital News Subscription Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Canada Employment Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Tuition Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Student Loan Interest Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Moving Expense Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Childcare Expense Deduction", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Carrying Charges Deduction", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Northern Residents Deduction", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Armed Forces Personnel Deduction", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Labour-Sponsored Investment Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Political Contribution Tax Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "First-Time Homebuyer Credit", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Home Buyers' Plan (HBP)", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Lifelong Learning Plan (LLP)", "institution_type": "Government", "typical_institutions": ["CRA"]},
        ]
    },
    "Government Programs & Entitlements": {
        "icon": "📋",
        "color": "#0984e3",
        "items": [
            {"name": "Employment Insurance (EI)", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "Ontario Works", "institution_type": "Government", "typical_institutions": ["Ontario Ministry of Community and Social Services"]},
            {"name": "ODSP — Ontario Disability Support Program", "institution_type": "Government", "typical_institutions": ["Ontario Ministry of Community and Social Services"]},
            {"name": "WSIB — Workers' Compensation", "institution_type": "Government", "typical_institutions": ["Workplace Safety and Insurance Board"]},
            {"name": "CPP Disability Benefit", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "CPP Retirement Benefit", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "CPP Survivor Benefit", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "CPP Children's Benefit", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "OAS Benefit", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "GIS Benefit", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "Allowance (Survivor)", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "Allowance (age 60-64)", "institution_type": "Government", "typical_institutions": ["Service Canada"]},
            {"name": "Veterans Benefits", "institution_type": "Government", "typical_institutions": ["Veterans Affairs Canada"]},
            {"name": "Veterans Independence Program", "institution_type": "Government", "typical_institutions": ["Veterans Affairs Canada"]},
            {"name": "War Veterans Allowance", "institution_type": "Government", "typical_institutions": ["Veterans Affairs Canada"]},
            {"name": "Indigenous Benefits", "institution_type": "Government", "typical_institutions": ["Indigenous Services Canada"]},
            {"name": "Status Card Benefits", "institution_type": "Government", "typical_institutions": ["Indigenous Services Canada"]},
            {"name": "Indian Residential School Settlement", "institution_type": "Government", "typical_institutions": ["IRSSA"]},
            {"name": "Jordan's Principle Benefits", "institution_type": "Government", "typical_institutions": ["Indigenous Services Canada"]},
        ]
    },
    "Loyalty Programs & Rewards": {
        "icon": "🎁",
        "color": "#fd79a8",
        "items": [
            {"name": "Aeroplan Points", "institution_type": "Airline", "typical_institutions": ["Air Canada"]},
            {"name": "Air Miles", "institution_type": "Loyalty", "typical_institutions": ["Air Miles"]},
            {"name": "Avios", "institution_type": "Airline", "typical_institutions": ["British Airways"]},
            {"name": "Marriott Bonvoy Points", "institution_type": "Hotel", "typical_institutions": ["Marriott"]},
            {"name": "Hilton Honors Points", "institution_type": "Hotel", "typical_institutions": ["Hilton"]},
            {"name": "IHG Rewards Points", "institution_type": "Hotel", "typical_institutions": ["IHG"]},
            {"name": "SPG Points", "institution_type": "Hotel", "typical_institutions": ["Marriott (legacy)"]},
            {"name": "WestJet Rewards", "institution_type": "Airline", "typical_institutions": ["WestJet"]},
            {"name": "Porter Airlines VIPorter", "institution_type": "Airline", "typical_institutions": ["Porter Airlines"]},
            {"name": "PC Optimum Points", "institution_type": "Retail", "typical_institutions": ["Loblaws", "Shoppers Drug Mart", "Esso"]},
            {"name": "Shoppers Drug Mart Points", "institution_type": "Retail", "typical_institutions": ["Shoppers Drug Mart"]},
            {"name": "HBC Rewards (The Bay)", "institution_type": "Retail", "typical_institutions": ["Hudson's Bay"]},
            {"name": "Canadian Tire Triangle Rewards", "institution_type": "Retail", "typical_institutions": ["Canadian Tire"]},
            {"name": "Petro-Points", "institution_type": "Gas", "typical_institutions": ["Petro-Canada"]},
            {"name": "Esso Extra", "institution_type": "Gas", "typical_institutions": ["Esso"]},
            {"name": "Sobeys/Affinity/Voilà Rewards", "institution_type": "Grocery", "typical_institutions": ["Sobeys"]},
            {"name": "Loblaws PC Points", "institution_type": "Grocery", "typical_institutions": ["Loblaws"]},
            {"name": "No Frills Rewards", "institution_type": "Grocery", "typical_institutions": ["No Frills"]},
            {"name": "Longos Rewards", "institution_type": "Grocery", "typical_institutions": ["Longos"]},
            {"name": "Food Basics Rewards", "institution_type": "Grocery", "typical_institutions": ["Food Basics"]},
            {"name": "Credit Card Travel Points", "institution_type": "Credit Card", "typical_institutions": ["Amex", "TD", "CIBC", "BMO", "Scotiabank"]},
            {"name": "Credit Card Cashback", "institution_type": "Credit Card", "typical_institutions": ["Various"]},
            {"name": "Credit Card Points (General)", "institution_type": "Credit Card", "typical_institutions": ["Various"]},
            {"name": "Gas Station Loyalty Points", "institution_type": "Gas", "typical_institutions": ["Various"]},
            {"name": "Airline Companion Passes", "institution_type": "Airline", "typical_institutions": ["WestJet", "Alaska Airlines"]},
            {"name": "Hotel Elite Status Benefits", "institution_type": "Hotel", "typical_institutions": ["Marriott", "Hilton", "IHG"]},
            {"name": "Priority Pass (Airport Lounge)", "institution_type": "Travel", "typical_institutions": ["Priority Pass"]},
            {"name": "NEXUS Card", "institution_type": "Government", "typical_institutions": ["CBSA", "CBP"]},
            {"name": "Global Entry", "institution_type": "Government", "typical_institutions": ["CBP"]},
            {"name": "TSA PreCheck", "institution_type": "Government", "typical_institutions": ["TSA"]},
            {"name": "CLEAR Membership", "institution_type": "Travel", "typical_institutions": ["CLEAR"]},
        ]
    },
    "Deposits & Security": {
        "icon": "🔒",
        "color": "#636e72",
        "items": [
            {"name": "Rental Security Deposit", "institution_type": "Landlord", "typical_institutions": []},
            {"name": "Utility Deposits", "institution_type": "Utility", "typical_institutions": ["Hydro One", "Toronto Hydro", "Enbridge"]},
            {"name": "Cell Phone Deposit", "institution_type": "Carrier", "typical_institutions": ["Rogers", "Bell", "Telus"]},
            {"name": "Landlord/Tenant Deposit", "institution_type": "Landlord", "typical_institutions": []},
            {"name": "Earnest Money (Real Estate)", "institution_type": "Real Estate", "typical_institutions": []},
            {"name": "Escrow Funds", "institution_type": "Financial", "typical_institutions": ["Title company"]},
            {"name": "Margin Account Deposit", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "IBKR"]},
            {"name": "Futures Margins", "institution_type": "Brokerage", "typical_institutions": ["IBKR"]},
            {"name": "Options Collateral", "institution_type": "Brokerage", "typical_institutions": ["IBKR", "Questrade"]},
            {"name": "Brokerage Cash Balance", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple", "IBKR"]},
            {"name": "Payroll Deductions (Refundable)", "institution_type": "Employer", "typical_institutions": ["Employer"]},
            {"name": "Tax Instalments Paid", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "GST/HST Collected But Not Remitted", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Deposit on Vehicle Purchase", "institution_type": "Dealer", "typical_institutions": ["Car dealership"]},
        ]
    },
    "Foreign Assets": {
        "icon": "🌍",
        "color": "#00cec9",
        "items": [
            {"name": "Foreign Bank Account", "institution_type": "Bank", "typical_institutions": ["US Banks", "International Banks"]},
            {"name": "Foreign Real Estate", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Foreign Investments", "institution_type": "Brokerage", "typical_institutions": ["IBKR", "Charles Schwab"]},
            {"name": "Foreign Pension (e.g., US 401k, IRA)", "institution_type": "Pension", "typical_institutions": ["US Employer"]},
            {"name": "Foreign Insurance", "institution_type": "Insurance", "typical_institutions": ["International insurers"]},
            {"name": "Foreign Business Interest", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Foreign Trust", "institution_type": "Trust", "typical_institutions": ["International trustees"]},
            {"name": "Foreign Life Insurance Cash Value", "institution_type": "Insurance", "typical_institutions": ["International insurers"]},
            {"name": "Foreign Bank Safety Deposit Box", "institution_type": "Bank", "typical_institutions": ["International banks"]},
        ]
    },
    "Trusts & Estates": {
        "icon": "📜",
        "color": "#fdcb6e",
        "items": [
            {"name": "Inter Vivos Trust (Living Trust)", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Testamentary Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Testamentary Spousal Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Alter Ego Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Joint Partner Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Qualified Disability Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Environmental Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Charitable Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Trust for Minors", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Trust for Incapacitated Persons", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Henson Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Spousal Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Family Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Bare Trust", "institution_type": "Trust", "typical_institutions": ["Trust company"]},
            {"name": "Constructive Trust Interest", "institution_type": "Legal", "typical_institutions": []},
        ]
    },
    "Joint & Shared Assets": {
        "icon": "🤝",
        "color": "#a29bfe",
        "items": [
            {"name": "Joint Bank Account", "institution_type": "Bank", "typical_institutions": ["TD", "RBC", "BMO"]},
            {"name": "Joint Investment Account", "institution_type": "Brokerage", "typical_institutions": ["Questrade", "Wealthsimple"]},
            {"name": "Joint Real Estate", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Joint Vehicle", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Joint Business Interest", "institution_type": "N/A", "typical_institutions": []},
            {"name": "Joint Credit Account", "institution_type": "Credit Card", "typical_institutions": ["TD", "RBC", "BMO"]},
            {"name": "Joint TFSA", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "Wealthsimple"]},
            {"name": "Joint RESP", "institution_type": "Bank/Brokerage", "typical_institutions": ["Questrade", "Wealthsimple"]},
            {"name": "Joint Crypto Wallet", "institution_type": "Exchange", "typical_institutions": ["Newton", "Shakepay"]},
        ]
    },
    "Contingent & Future Interests": {
        "icon": "⏳",
        "color": "#636e72",
        "items": [
            {"name": "Remainder Interest in Trust", "institution_type": "Trust", "typical_institutions": []},
            {"name": "Reversionary Interest", "institution_type": "Trust", "typical_institutions": []},
            {"name": "Power of Appointment", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Survivorship Interest", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Vested Remainder", "institution_type": "Trust", "typical_institutions": []},
            {"name": "Contingent Remainder", "institution_type": "Trust", "typical_institutions": []},
            {"name": "Life Estate", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Expectant Estate (Future Inheritance)", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Rights of First Refusal", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Option to Purchase", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Beneficiary Interest in Estate", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Future Inheritance Rights", "institution_type": "Legal", "typical_institutions": []},
        ]
    },
    "Miscellaneous Assets": {
        "icon": "📦",
        "color": "#74b9ff",
        "items": [
            {"name": "Gift Cards", "institution_type": "Retail", "typical_institutions": ["Various"]},
            {"name": "Vouchers", "institution_type": "Retail", "typical_institutions": ["Various"]},
            {"name": "Prepaid Services", "institution_type": "Various", "typical_institutions": ["Various"]},
            {"name": "Store Credits", "institution_type": "Retail", "typical_institutions": ["Various"]},
            {"name": "Return Refunds Pending", "institution_type": "Retail", "typical_institutions": ["Various"]},
            {"name": "Tax Refund Pending", "institution_type": "Government", "typical_institutions": ["CRA"]},
            {"name": "Rebates Pending", "institution_type": "Manufacturer", "typical_institutions": ["Various"]},
            {"name": "Winning Lottery Ticket (Unclaimed)", "institution_type": "Government", "typical_institutions": ["OLG"]},
            {"name": "Contest Prize Pending", "institution_type": "Various", "typical_institutions": ["Various"]},
            {"name": "Judgment Pending (Legal)", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Lawsuit Claim Pending", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Insurance Claim Pending", "institution_type": "Insurance", "typical_institutions": ["Various"]},
            {"name": "WSIB Claim Pending", "institution_type": "Government", "typical_institutions": ["WSIB"]},
            {"name": "Class Action Settlement Pending", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Vendor Rebates", "institution_type": "Manufacturer", "typical_institutions": ["Various"]},
            {"name": "Manufacturer Rebates", "institution_type": "Manufacturer", "typical_institutions": ["Various"]},
            {"name": "Rain Checks", "institution_type": "Retail", "typical_institutions": ["Various"]},
        ]
    },
    "Water, Mineral & Land Rights": {
        "icon": "⛏️",
        "color": "#00b894",
        "items": [
            {"name": "Water Rights", "institution_type": "Government", "typical_institutions": ["Ontario Ministry of Natural Resources"]},
            {"name": "Mineral Rights", "institution_type": "Government", "typical_institutions": ["Ontario Ministry of Mines"]},
            {"name": "Timber Rights", "institution_type": "Government", "typical_institutions": ["Ontario Ministry of Natural Resources"]},
            {"name": "Hunting Rights", "institution_type": "Government", "typical_institutions": ["Ontario MNRF"]},
            {"name": "Fishing Rights", "institution_type": "Government", "typical_institutions": ["Ontario MNRF"]},
            {"name": "Grazing Rights", "institution_type": "Government", "typical_institutions": ["Ontario Ministry of Agriculture"]},
            {"name": "Easements", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Rights of Way", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Air Rights", "institution_type": "Legal", "typical_institutions": []},
            {"name": "Development Rights", "institution_type": "Government", "typical_institutions": ["Municipal planning"]},
            {"name": "Zoning Variance", "institution_type": "Government", "typical_institutions": ["Municipal planning"]},
            {"name": "Building Permit (Transferable)", "institution_type": "Government", "typical_institutions": ["Municipal building dept"]},
            {"name": "Special Use Permit", "institution_type": "Government", "typical_institutions": ["Municipal planning"]},
            {"name": "Conservation Easement", "institution_type": "Legal", "typical_institutions": ["Conservation authority"]},
            {"name": "Covenant Rights", "institution_type": "Legal", "typical_institutions": []},
        ]
    },
}

# =============================================================================
# TIER CONFIGURATION (versioning_plan.md)
# =============================================================================

# Tiers: free | family | planning | advisor (advisor deferred)
FREE_CATEGORIES = [
    "Cash & Cash Equivalents", "Fixed Income Investments",
    "Equities & Investment Funds", "Registered Accounts — Canada",
    "Pension Benefits", "Insurance Products", "Real Estate",
    "Vehicles & Transportation", "Personal Property — Valuables",
    "Household & Electronics", "Loyalty Programs & Rewards",
    "Deposits & Security", "Joint & Shared Assets",
    "Government Benefits & Tax Credits", "Government Programs & Entitlements",
]

# Family adds 5 categories (Employment + Digital Online/Business/Content/Accounts)
FAMILY_CATEGORIES = FREE_CATEGORIES + [
    "Employment & Compensation Assets",
    "Digital Assets — Online Presence",
    "Digital Assets — Online Businesses & Income",
    "Digital Assets — Content & IP",
    "Digital Assets — Accounts & Subscriptions",
]

TIER_CATEGORIES = {
    "free": FREE_CATEGORIES,
    "family": FAMILY_CATEGORIES,
    "planning": list(ASSET_CATEGORIES.keys()),  # full 32-category catalog
    "advisor": list(ASSET_CATEGORIES.keys()),
}

# =============================================================================
# FIELD DEFINITIONS (108 fields)
# =============================================================================

FIELD_DEFINITIONS = {
    # Core Identity
    "id": {"group": "Core Identity", "type": "text", "label": "ID"},
    "category": {"group": "Core Identity", "type": "text", "label": "Category"},
    "subcategory": {"group": "Core Identity", "type": "text", "label": "Subcategory"},
    "asset_name": {"group": "Core Identity", "type": "text", "label": "Asset Name"},
    "asset_type": {"group": "Core Identity", "type": "text", "label": "Asset Type"},
    
    # Ownership
    "owner": {"group": "Ownership", "type": "text", "label": "Owner"},
    "joint_owner": {"group": "Ownership", "type": "text", "label": "Joint Owner"},
    "beneficiary": {"group": "Ownership", "type": "text", "label": "Beneficiary"},
    "beneficiary_type": {"group": "Ownership", "type": "select", "label": "Beneficiary Type", "options": ["", "Spouse", "Child", "Children equally", "Estate", "Trust", "Other"]},
    "custodian": {"group": "Ownership", "type": "text", "label": "Custodian"},
    "nominee": {"group": "Ownership", "type": "text", "label": "Nominee"},
    "trust_name": {"group": "Ownership", "type": "text", "label": "Trust Name"},
    "corporation": {"group": "Ownership", "type": "text", "label": "Corporation"},
    
    # Institution & Access
    "institution": {"group": "Institution & Access", "type": "text", "label": "Institution"},
    "institution_type": {"group": "Institution & Access", "type": "text", "label": "Institution Type"},
    "branch": {"group": "Institution & Access", "type": "text", "label": "Branch"},
    "account_number": {"group": "Institution & Access", "type": "text", "label": "Account #"},
    "login_url": {"group": "Institution & Access", "type": "url", "label": "Login URL"},
    "login_username": {"group": "Institution & Access", "type": "text", "label": "Username"},
    "login_password": {"group": "Institution & Access", "type": "password", "label": "Password"},
    "two_factor": {"group": "Institution & Access", "type": "text", "label": "2FA Method"},
    "security_questions": {"group": "Institution & Access", "type": "text", "label": "Security Q&A Location"},
    
    # Financial Value
    "currency": {"group": "Financial Value", "type": "select", "label": "Currency", "options": ["CAD", "USD", "EUR", "GBP", "BTC", "ETH", "Other"]},
    "acb": {"group": "Financial Value", "type": "currency", "label": "ACB"},
    "acb_usd": {"group": "Financial Value", "type": "currency", "label": "ACB (USD)"},
    "fmv": {"group": "Financial Value", "type": "currency", "label": "FMV"},
    "fmv_usd": {"group": "Financial Value", "type": "currency", "label": "FMV (USD)"},
    "cost_basis": {"group": "Financial Value", "type": "currency", "label": "Cost Basis"},
    "purchase_price": {"group": "Financial Value", "type": "currency", "label": "Purchase Price"},
    "current_balance": {"group": "Financial Value", "type": "currency", "label": "Current Balance"},
    "market_value": {"group": "Financial Value", "type": "currency", "label": "Market Value"},
    "equity": {"group": "Financial Value", "type": "currency", "label": "Equity"},
    "unrealized_gain": {"group": "Financial Value", "type": "currency", "label": "Unrealized Gain"},
    "unrealized_gain_pct": {"group": "Financial Value", "type": "percent", "label": "Unrealized Gain %"},
    "annual_income": {"group": "Financial Value", "type": "currency", "label": "Annual Income"},
    "yield_pct": {"group": "Financial Value", "type": "percent", "label": "Yield %"},
    "interest_rate": {"group": "Financial Value", "type": "percent", "label": "Interest Rate"},
    "dividend_rate": {"group": "Financial Value", "type": "text", "label": "Dividend Rate"},
    
    # Date Fields
    "open_date": {"group": "Dates", "type": "date", "label": "Open Date"},
    "maturity_date": {"group": "Dates", "type": "date", "label": "Maturity Date"},
    "purchase_date": {"group": "Dates", "type": "date", "label": "Purchase Date"},
    "acquisition_date": {"group": "Dates", "type": "date", "label": "Acquisition Date"},
    "inception_date": {"group": "Dates", "type": "date", "label": "Inception Date"},
    "expiry_date": {"group": "Dates", "type": "date", "label": "Expiry Date"},
    "last_valuation": {"group": "Dates", "type": "date", "label": "Last Valuation"},
    "last_update": {"group": "Dates", "type": "date", "label": "Last Updated"},
    "next_review": {"group": "Dates", "type": "date", "label": "Next Review"},
    "transfer_date": {"group": "Dates", "type": "date", "label": "Transfer Date"},
    
    # Registration & Tax
    "registration": {"group": "Registration & Tax", "type": "select", "label": "Registration", "options": ["", "TFSA", "RRSP", "RESP", "RRIF", "LIRA", "LIF", "LRIF", "PRIF", "RPP", "DPSP", "RDSP", "FHSA", "Non-registered", "Margin", "Other"]},
    "tax_treatment": {"group": "Registration & Tax", "type": "select", "label": "Tax Treatment", "options": ["", "Tax-free", "Tax-deferred", "Taxable", "Capital gains", "Other"]},
    "contribution_room": {"group": "Registration & Tax", "type": "currency", "label": "Contribution Room"},
    "contributions_ytd": {"group": "Registration & Tax", "type": "currency", "label": "Contributions YTD"},
    "withdrawals_ytd": {"group": "Registration & Tax", "type": "currency", "label": "Withdrawals YTD"},
    "rrsp_deduction": {"group": "Registration & Tax", "type": "currency", "label": "RRSP Deduction"},
    "cesg": {"group": "Registration & Tax", "type": "currency", "label": "CESG"},
    "clb": {"group": "Registration & Tax", "type": "currency", "label": "CLB"},
    "provincial_grant": {"group": "Registration & Tax", "type": "currency", "label": "Provincial Grant"},
    "foreign_tax_credit": {"group": "Registration & Tax", "type": "currency", "label": "Foreign Tax Credit"},
    "withholding_tax": {"group": "Registration & Tax", "type": "currency", "label": "Withholding Tax"},
    "asset_allocation": {"group": "Registration & Tax", "type": "text", "label": "Asset Allocation"},
    
    # Location & Access
    "physical_location": {"group": "Location & Access", "type": "text", "label": "Physical Location"},
    "safe_deposit_box": {"group": "Location & Access", "type": "text", "label": "Safe Deposit Box"},
    "digital_wallet": {"group": "Location & Access", "type": "text", "label": "Digital Wallet"},
    "exchange": {"group": "Location & Access", "type": "text", "label": "Exchange"},
    "online_access_url": {"group": "Location & Access", "type": "url", "label": "Online Access URL"},
    "support_contact": {"group": "Location & Access", "type": "text", "label": "Support Contact"},
    "advisor_name": {"group": "Location & Access", "type": "text", "label": "Advisor Name"},
    "advisor_contact": {"group": "Location & Access", "type": "text", "label": "Advisor Contact"},
    
    # Insurance & Protection
    "insurance_coverage": {"group": "Insurance & Protection", "type": "currency", "label": "Insurance Coverage"},
    "insurance_provider": {"group": "Insurance & Protection", "type": "text", "label": "Insurance Provider"},
    "insurance_policy": {"group": "Insurance & Protection", "type": "text", "label": "Policy #"},
    "insurance_premium": {"group": "Insurance & Protection", "type": "currency", "label": "Annual Premium"},
    "insured_value": {"group": "Insurance & Protection", "type": "currency", "label": "Insured Value"},
    "replacement_cost": {"group": "Insurance & Protection", "type": "currency", "label": "Replacement Cost"},
    
    # Status & Control
    "status": {"group": "Status & Control", "type": "select", "label": "Status", "options": ["Active", "Dormant", "Pending", "Closed"]},
    "priority": {"group": "Status & Control", "type": "select", "label": "Priority", "options": ["High", "Medium", "Low"]},
    "volatility": {"group": "Status & Control", "type": "select", "label": "Volatility", "options": ["", "Low", "Medium", "High"]},
    "liquidity": {"group": "Status & Control", "type": "select", "label": "Liquidity", "options": ["", "High", "Medium", "Low"]},
    "transferable": {"group": "Status & Control", "type": "select", "label": "Transferable", "options": ["Yes", "No"]},
    "pledgeable": {"group": "Status & Control", "type": "select", "label": "Pledgeable", "options": ["Yes", "No"]},
    "accessible": {"group": "Status & Control", "type": "select", "label": "Accessible", "options": ["", "None", "Probate", "LIRA lock-in", "Other"]},
    
    # Beneficiary Designation
    "primary_beneficiary": {"group": "Beneficiary Designation", "type": "text", "label": "Primary Beneficiary"},
    "contingent_beneficiary": {"group": "Beneficiary Designation", "type": "text", "label": "Contingent Beneficiary"},
    "beneficiary_pct_primary": {"group": "Beneficiary Designation", "type": "percent", "label": "Primary %"},
    "beneficiary_pct_contingent": {"group": "Beneficiary Designation", "type": "percent", "label": "Contingent %"},
    "pod": {"group": "Beneficiary Designation", "type": "select", "label": "POD", "options": ["Yes", "No"]},
    "tod": {"group": "Beneficiary Designation", "type": "select", "label": "TOD", "options": ["Yes", "No"]},
    "poa": {"group": "Beneficiary Designation", "type": "text", "label": "POA Holder"},
    "mandate": {"group": "Beneficiary Designation", "type": "text", "label": "Mandate (Quebec)"},
    
    # Estate Planning
    "probate_excluded": {"group": "Estate Planning", "type": "select", "label": "Probate Excluded", "options": ["Yes", "No"]},
    "will_clause": {"group": "Estate Planning", "type": "text", "label": "Will Clause"},
    "trust_clause": {"group": "Estate Planning", "type": "text", "label": "Trust Clause"},
    "estate_duty": {"group": "Estate Planning", "type": "select", "label": "Estate Duty", "options": ["Yes", "No"]},
    "capital_gains_exemption": {"group": "Estate Planning", "type": "select", "label": "LCGE Eligible", "options": ["Yes", "No", "QSBCS"]},
    "succession_plan": {"group": "Estate Planning", "type": "text", "label": "Succession Plan"},
    "power_of_appointment": {"group": "Estate Planning", "type": "text", "label": "POA for Assets"},
    
    # Documentation
    "document_path": {"group": "Documentation", "type": "text", "label": "Document Path"},
    "document_reference": {"group": "Documentation", "type": "text", "label": "Document Ref #"},
    "last_statement": {"group": "Documentation", "type": "date", "label": "Last Statement"},
    "statement_frequency": {"group": "Documentation", "type": "select", "label": "Statement Freq", "options": ["", "Monthly", "Quarterly", "Annually"]},
    "tax_slip_type": {"group": "Documentation", "type": "select", "label": "Tax Slip Type", "options": ["", "T5", "T3", "T5008", "T4A", "T4RIF", "T4RSP", "T101", "Other"]},
    "tax_slip_received": {"group": "Documentation", "type": "select", "label": "Tax Slip Received", "options": ["Yes", "No", "N/A"]},
    "annual_report": {"group": "Documentation", "type": "select", "label": "Annual Report", "options": ["Yes", "No", "N/A"]},
    
    # Notes
    "notes": {"group": "Notes", "type": "textarea", "label": "Notes"},
    "alert": {"group": "Notes", "type": "text", "label": "Alert"},
    "todo": {"group": "Notes", "type": "text", "label": "To-Do"},
    "last_modified_by": {"group": "Notes", "type": "text", "label": "Last Modified By"},
    "source": {"group": "Notes", "type": "select", "label": "Source", "options": ["", "Purchase", "Inheritance", "Gift", "Employment", "Government", "Created", "Other"]},
}

# =============================================================================
# GENERATE ASSET DATA
# =============================================================================

def generate_all_assets(tier: str = "planning") -> list[dict]:
    """Generate all assets for the given tier with default values."""
    allowed = TIER_CATEGORIES.get(tier, TIER_CATEGORIES["planning"])
    assets = []
    asset_id = 1
    
    for category_name, category_data in ASSET_CATEGORIES.items():
        if category_name not in allowed:
            continue
        icon = category_data["icon"]
        color = category_data["color"]
        
        for item in category_data["items"]:
            asset = {
                "id": f"A-{asset_id:04d}",
                "category": category_name,
                "subcategory": item["name"],
                "asset_name": item["name"],
                "asset_type": "Financial" if category_name in [
                    "Cash & Cash Equivalents", "Fixed Income Investments", 
                    "Equities & Investment Funds", "Registered Accounts — Canada",
                    "Cryptocurrencies", "Stablecoins", "DeFi & Staking"
                ] else "Real Estate" if category_name == "Real Estate" else 
                "Digital" if "Digital" in category_name or "NFT" in category_name else 
                "Physical" if category_name in ["Vehicles & Transportation", "Personal Property — Valuables", "Household & Electronics"] else
                "Legal" if category_name in ["Intellectual Property", "Trusts & Estates", "Contingent & Future Interests"] else
                "Government" if "Government" in category_name else "Other",
                
                # Institution defaults
                "institution": item.get("typical_institutions", [""])[0] if item.get("typical_institutions") else "",
                "institution_type": item.get("institution_type", ""),
                
                # Status defaults
                "status": "Active",
                "priority": "Medium",
                "currency": "CAD",
                
                # Financial defaults
                "acb": 0.0,
                "fmv": 0.0,
                "current_balance": 0.0,
                
                # Date defaults
                "last_update": date.today().isoformat(),
                
                # Tax defaults
                "tax_treatment": item.get("tax_treatment", ""),
                "registration": "",
                
                # Empty fields (to be filled by user)
                "owner": "",
                "joint_owner": "",
                "beneficiary": "",
                "beneficiary_type": "",
                "custodian": "",
                "nominee": "",
                "trust_name": "",
                "corporation": "",
                "branch": "",
                "account_number": "",
                "login_url": "",
                "login_username": "",
                "login_password": "",
                "two_factor": "",
                "security_questions": "",
                "acb_usd": 0.0,
                "fmv_usd": 0.0,
                "cost_basis": 0.0,
                "purchase_price": 0.0,
                "market_value": 0.0,
                "equity": 0.0,
                "unrealized_gain": 0.0,
                "unrealized_gain_pct": 0.0,
                "annual_income": 0.0,
                "yield_pct": 0.0,
                "interest_rate": 0.0,
                "dividend_rate": "",
                "open_date": "",
                "maturity_date": "",
                "purchase_date": "",
                "acquisition_date": "",
                "inception_date": "",
                "expiry_date": "",
                "last_valuation": "",
                "next_review": "",
                "transfer_date": "",
                "contribution_room": 0.0,
                "contributions_ytd": 0.0,
                "withdrawals_ytd": 0.0,
                "rrsp_deduction": 0.0,
                "cesg": 0.0,
                "clb": 0.0,
                "provincial_grant": 0.0,
                "foreign_tax_credit": 0.0,
                "withholding_tax": 0.0,
                "asset_allocation": "",
                "physical_location": "",
                "safe_deposit_box": "",
                "digital_wallet": "",
                "exchange": "",
                "online_access_url": "",
                "support_contact": "",
                "advisor_name": "",
                "advisor_contact": "",
                "insurance_coverage": 0.0,
                "insurance_provider": "",
                "insurance_policy": "",
                "insurance_premium": 0.0,
                "insured_value": 0.0,
                "replacement_cost": 0.0,
                "volatility": "",
                "liquidity": "",
                "transferable": "",
                "pledgeable": "",
                "accessible": "",
                "primary_beneficiary": "",
                "contingent_beneficiary": "",
                "beneficiary_pct_primary": 0.0,
                "beneficiary_pct_contingent": 0.0,
                "pod": "",
                "tod": "",
                "poa": "",
                "mandate": "",
                "probate_excluded": "",
                "will_clause": "",
                "trust_clause": "",
                "estate_duty": "",
                "capital_gains_exemption": "",
                "succession_plan": "",
                "power_of_appointment": "",
                "document_path": "",
                "document_reference": "",
                "last_statement": "",
                "statement_frequency": "",
                "tax_slip_type": "",
                "tax_slip_received": "",
                "annual_report": "",
                "notes": "",
                "alert": "",
                "todo": "",
                "last_modified_by": "",
                "source": "",
            }
            assets.append(asset)
            asset_id += 1
    
    return assets

# =============================================================================
# DEMO FIXTURE
# =============================================================================

def _demo_find(assets: list[dict], subcategory: str) -> Optional[dict]:
    """Return the first asset whose subcategory matches (template asset)."""
    for a in assets:
        if a.get("subcategory") == subcategory:
            return a
    return None


def populate_demo_data(assets: list[dict]) -> list[dict]:
    """Overlay a realistic Ontario family fixture on the blank template.

    Two adults (James & Mei Chen), two children (RESP beneficiaries), a family
    home with mortgage equity, TFSAs/RRSPs/RESP, a dormant crypto wallet, and a
    mixed set of complete/blank assets so charts, audit, and search have
    meaningful data. Only fields present in the 108-field schema are set.
    """
    def fill(subcategory, **values):
        asset = _demo_find(assets, subcategory)
        if not asset:
            return None
        asset.update(values)
        asset["last_modified_by"] = "demo-fixture"
        asset["source"] = "demo"
        return asset

    # --- Family: James Chen (owner), Mei Chen (spouse), Ethan (12), Olivia (8)
    # --- Cash & cash equivalents
    fill("Chequing Account", fmv=12400.0, acb=12400.0, owner="James Chen",
         institution="TD Canada Trust", account_number="64-5128", last_update="2026-07-15")
    fill("High-Interest Savings Account", fmv=45500.0, acb=45500.0,
         owner="James Chen", joint_owner="Mei Chen", institution="EQ Bank",
         account_number="1004-2291", interest_rate=3.1, annual_income=1410.0,
         last_update="2026-07-15")
    fill("Money Market Account", fmv=18000.0, acb=18000.0, owner="James Chen",
         institution="Wealthsimple", interest_rate=2.4, annual_income=432.0,
         last_update="2026-06-20")
    fill("Cashable GIC", fmv=15000.0, acb=15000.0, owner="James Chen",
         institution="EQ Bank", maturity_date="2027-03-01", interest_rate=4.1,
         last_update="2026-03-01")

    # --- Fixed income
    fill("Guaranteed Investment Certificates (GICs)", fmv=30000.0, acb=30000.0,
         owner="Mei Chen", institution="Oaken Financial", maturity_date="2027-06-30",
         interest_rate=4.5, last_update="2026-06-30")
    fill("Government of Canada Bonds", fmv=12000.0, acb=11000.0, owner="James Chen",
         institution="Government of Canada", maturity_date="2031-11-01",
         interest_rate=3.2, last_update="2026-01-12")

    # --- Equities & funds
    fill("Canadian Equity ETFs", fmv=62500.0, acb=45500.0, owner="James Chen",
         joint_owner="Mei Chen", institution="Questrade", registration="Non-registered",
         annual_income=980.0, dividend_rate="1.6%", unrealized_gain=17000.0,
         unrealized_gain_pct=37.4, volatility="Medium", liquidity="High",
         last_update="2026-07-10")
    fill("US Equity ETFs", fmv=34800.0, acb=30000.0, owner="James Chen",
         institution="Questrade", registration="RRSP", annual_income=320.0,
         volatility="Medium", liquidity="High", last_update="2026-07-10")
    fill("Individual Canadian Stocks (TSX)", fmv=25000.0, acb=19800.0,
         owner="James Chen", institution="TD Direct Investing",
         registration="Non-registered", annual_income=415.0, volatility="High",
         liquidity="High", last_update="2026-05-22")

    # --- Registered accounts
    fill("TFSA (Tax-Free Savings Account)", fmv=78500.0, acb=70000.0,
         owner="James Chen", institution="Wealthsimple", registration="TFSA",
         contribution_room=6500.0, contributions_ytd=8200.0,  # demo over-contribution (audit red)
         primary_beneficiary="Mei Chen", beneficiary_pct_primary=100.0,
         probate_excluded="Yes", asset_allocation="60% equities / 40% fixed",
         volatility="Medium", liquidity="High", last_update="2026-07-01")
    fill("RRSP (Registered Retirement Savings Plan)", fmv=210000.0, acb=175000.0,
         owner="James Chen", institution="TD Direct Investing", registration="RRSP",
         contribution_room=23000.0, contributions_ytd=6000.0, rrsp_deduction=6000.0,
         primary_beneficiary="Mei Chen", beneficiary_pct_primary=100.0,
         probate_excluded="Yes", asset_allocation="55% equities / 45% fixed",
         volatility="Medium", liquidity="Low", last_update="2026-06-25")
    fill("RESP (Registered Education Savings Plan)", fmv=52000.0, acb=42000.0,
         owner="James Chen", institution="Wealthsimple", registration="RESP",
         contributions_ytd=2500.0, cesg=7200.0, clb=500.0,
         primary_beneficiary="Ethan Chen", beneficiary_pct_primary=60.0,
         contingent_beneficiary="Olivia Chen", beneficiary_pct_contingent=40.0,
         probate_excluded="Yes", asset_allocation="70% equities / 30% fixed",
         volatility="Medium", liquidity="Medium", last_update="2026-07-05",
         notes="Education fund for **Ethan (age 12)** and *Olivia (age 8)*.\n\n[Wealthsimple RESP](https://www.wealthsimple.com) — family plan, both children named.",
         todo="Contribute $2,500 before Dec 31 to maximize CESG match.")
    fill("FHSA (First Home Savings Account)", fmv=8000.0, acb=8000.0,
         owner="James Chen", institution="EQ Bank", registration="FHSA",
         contributions_ytd=1000.0, contribution_room=7000.0, probate_excluded="Yes",
         liquidity="High", last_update="2026-02-14")
    fill("RRIF (Registered Retirement Income Fund)", fmv=64000.0, acb=60000.0,
         owner="Mei Chen", institution="RBC", registration="RRIF",
         primary_beneficiary="James Chen", beneficiary_pct_primary=100.0,
         probate_excluded="Yes", volatility="Low", liquidity="Low",
         last_update="2025-11-18")  # stale (audit yellow)

    # --- Real estate
    fill("Primary Residence", fmv=1250000.0, purchase_price=880000.0,
         market_value=1250000.0, equity=420000.0, owner="James Chen",
         joint_owner="Mei Chen", institution="—", registration="Primary Residence",
         physical_location="42 Maplewood Cres, Markham, ON",
         probate_excluded="No", pod="Joint Tenancy",
         last_valuation="2026-01-10", last_update="2026-07-12",
         notes="Family home purchased **2019**. Title held as *joint tenants* — bypasses probate on first death.")
    fill("Rental Property (Residential)", fmv=620000.0, purchase_price=480000.0,
         market_value=620000.0, equity=145000.0, owner="James Chen",
         institution="—", annual_income=26400.0, physical_location="18 Birch Ave, Scarborough, ON",
         last_valuation="2026-01-15", last_update="2026-07-12",
         notes="Tenant since 2022. Monthly rent **$2,200**. See lease + Ontario Standard Lease on file.",
         todo="Notify tenant of 2026 rent guideline increase (2.5%) by Nov 1.")

    # --- Vehicles
    fill("Car/Truck", fmv=28500.0, acb=36000.0, owner="Mei Chen",
         institution="Honda Canada Finance", purchase_date="2022-05-01",
         last_update="2026-04-30", notes="2022 Honda CR-V — financed, ~$9,000 remaining.")

    # --- Insurance
    fill("Term Life Insurance", insurance_coverage=500000.0, insured_value=500000.0,
         insurance_provider="Manulife", insurance_policy="T12345678",
         insurance_premium=62.5, owner="James Chen", primary_beneficiary="Mei Chen",
         beneficiary_pct_primary=100.0, probate_excluded="Yes",
         last_update="2026-01-08", notes="20-year term, **$500K**, started 2021.",
         alert="Renewal in 2031 — re-quote in 2030.")
    fill("Whole Life Insurance", insurance_coverage=250000.0, insured_value=250000.0,
         insurance_provider="Sun Life", insurance_policy="W87654321",
         insurance_premium=180.0, owner="Mei Chen", primary_beneficiary="James Chen",
         beneficiary_pct_primary=100.0, probate_excluded="Yes",
         last_update="2024-09-30")  # stale (audit yellow)

    # --- Crypto (dormant wallet)
    fill("Bitcoin (BTC)", fmv=8500.0, acb=12400.0, owner="James Chen",
         status="Dormant", exchange="Cold storage (hardware wallet)",
         digital_wallet="Ledger Nano X", volatility="Very High", liquidity="High",
         last_update="2025-12-01", notes="Held on **Ledger Nano X** — seed phrase location documented, **not stored digitally**.",
         todo="Review custody protocol annually; update location card.")
    fill("Ethereum (ETH)", fmv=3200.0, acb=4500.0, owner="James Chen",
         status="Dormant", exchange="Cold storage (hardware wallet)",
         digital_wallet="Ledger Nano X", volatility="Very High", liquidity="High",
         last_update="2025-12-01")
    fill("Hardware Wallet (Ledger, Trezor, etc.)", fmv=150.0, owner="James Chen",
         status="Active", digital_wallet="Ledger Nano X",
         physical_location="Bank safety deposit box #1442",
         last_update="2025-12-01", notes="Recovery card locations: lawyer office + safety deposit box (2-of-3 shares).")

    # --- Pension & benefits
    fill("Employer Defined-Benefit Pension", fmv=340000.0, owner="James Chen",
         institution="Ontario Teachers' Pension Plan", annual_income=18000.0,
         probate_excluded="Yes", last_update="2026-06-01",
         notes="DB pension — projected annual pension at 65: **$38,000**.")
    fill("CPP Retirement Benefit", fmv=90000.0, owner="James Chen",
         institution="Government of Canada", annual_income=0.0, last_update="2026-01-01")
    fill("OAS — Old Age Security Credits", fmv=40000.0, owner="Mei Chen",
         institution="Government of Canada", annual_income=0.0, last_update="2026-01-01")

    # --- Valuables
    fill("Jewelry", fmv=15000.0, owner="Mei Chen", physical_location="Home safe",
         last_update="2026-03-01", notes="Engagement ring + heirloom pieces. Appraised 2025.")
    fill("Artwork (Paintings, Sculptures)", fmv=12000.0, owner="James Chen",
         physical_location="Living room", last_update="2026-03-01")

    return assets


# =============================================================================
# MARKDOWN GENERATOR
# =============================================================================

def generate_markdown(assets: list[dict], output_path: str, lang: str = "en"):
    """Generate formatted Markdown file."""
    
    # Group by category
    categories = {}
    for asset in assets:
        cat = asset["category"]
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(asset)
    
    # Calculate totals
    total_assets = len(assets)
    total_fmv = sum(a.get("fmv", 0) or 0 for a in assets)
    total_income = sum(a.get("annual_income", 0) or 0 for a in assets)
    
    md = []
    md.append(f"# {_('md_title', lang)} — Canadian Family in Ontario")
    md.append(f"\n**{_('md_generated', lang)}:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    md.append(f"**{_('total_assets', lang)}:** {total_assets} | **{_('total_fmv', lang)}:** ${total_fmv:,.2f} | **{_('annual_income', lang)}:** ${total_income:,.2f}")
    md.append("")
    md.append("---")
    md.append("")
    
    # Summary table
    md.append(f"## {_('sheet_summary', lang)}")
    md.append("")
    md.append(f"| {_('md_category', lang)} | {_('total_assets', lang)} | {_('total_fmv', lang)} | {_('annual_income', lang)} |")
    md.append("|----------|-------|-----------|---------------|")
    
    for cat_name in categories:
        cat_assets = categories[cat_name]
        cat_fmv = sum(a.get("fmv", 0) or 0 for a in cat_assets)
        cat_income = sum(a.get("annual_income", 0) or 0 for a in cat_assets)
        md.append(f"| {cat_name} | {len(cat_assets)} | ${cat_fmv:,.2f} | ${cat_income:,.2f} |")
    
    md.append(f"| **TOTAL** | **{total_assets}** | **${total_fmv:,.2f}** | **${total_income:,.2f}** |")
    md.append("")
    md.append("---")
    md.append("")
    
    # Detailed sections
    for cat_name in categories:
        cat_assets = categories[cat_name]
        # Look up icon using the original English category name if translation happened
        source_cat = next((k for k, v in CATEGORY_TRANSLATIONS.get(lang, {}).items() if v == cat_name), cat_name)
        cat_data = ASSET_CATEGORIES.get(source_cat, {})
        icon = cat_data.get("icon", "📋")
        
        md.append(f"## {icon} {cat_name}")
        md.append(f"*{len(cat_assets)} {_('total_assets', lang).lower()}*")
        md.append("")
        
        # Table header
        md.append(f"| {_('md_id', lang)} | {_('md_asset', lang)} | {_('md_owner', lang)} | {_('md_institution', lang)} | {_('col_account_number', lang)} | {_('md_value', lang)} | {_('md_status', lang)} |")
        md.append("|-----|------------|-------|-------------|-----------|-----|--------|")
        
        for asset in cat_assets:
            fmv = asset.get("fmv", 0) or 0
            fmv_str = f"${fmv:,.2f}" if fmv > 0 else "-"
            account = asset.get("account_number", "") or "-"
            owner = asset.get("owner", "") or "-"
            institution = asset.get("institution", "") or "-"
            status = asset.get("status", "Active")
            
            md.append(f"| {asset['id']} | {asset['asset_name']} | {owner} | {institution} | {account} | {fmv_str} | {status} |")
        
        md.append("")
        md.append("---")
        md.append("")
    
    # Write file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(md))
    
    print(f"[OK] Markdown generated: {output_path}")

# =============================================================================
# EXCEL GENERATOR
# =============================================================================

def generate_excel(assets: list[dict], output_path: str, lang: str = "en"):
    """Generate Excel workbook with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[SKIP] openpyxl not installed. Run: pip install openpyxl")
        return
    
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    currency_format = '#,##0.00'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
    
    # Sheet 1: All Assets
    ws1 = wb.active
    ws1.title = _("sheet_all_assets", lang)
    
    headers = [
        _("col_id", lang), _("col_category", lang), _("col_asset_name", lang),
        _("col_owner", lang), _("col_joint_owner", lang), _("col_beneficiary", lang),
        _("col_institution", lang), _("col_account_number", lang), _("col_registration", lang),
        _("col_currency", lang), _("col_acb", lang), _("col_fmv", lang),
        _("col_annual_income", lang), _("col_status", lang), _("col_priority", lang),
        _("col_last_update", lang), _("col_notes", lang)
    ]
    
    ws1.append(headers)
    style_header(ws1)
    
    for asset in assets:
        ws1.append([
            asset.get("id", ""),
            asset.get("category", ""),
            asset.get("asset_name", ""),
            asset.get("owner", ""),
            asset.get("joint_owner", ""),
            asset.get("beneficiary", ""),
            asset.get("institution", ""),
            asset.get("account_number", ""),
            asset.get("registration", ""),
            asset.get("currency", "CAD"),
            asset.get("acb", 0),
            asset.get("fmv", 0),
            asset.get("annual_income", 0),
            asset.get("status", ""),
            asset.get("priority", ""),
            asset.get("last_update", ""),
            asset.get("notes", ""),
        ])
    
    # Format currency columns
    for row in ws1.iter_rows(min_row=2, min_col=11, max_col=13):
        for cell in row:
            cell.number_format = currency_format
    
    # Auto-width
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 18
    
    # Sheet 2: By Category
    ws2 = wb.create_sheet(_("sheet_by_category", lang))
    ws2.append([_("col_category", lang), _("summary_total_assets", lang), _("summary_total_fmv", lang), _("summary_total_income", lang)])
    style_header(ws2)
    
    categories = {}
    for asset in assets:
        cat = asset["category"]
        if cat not in categories:
            categories[cat] = {"count": 0, "fmv": 0, "income": 0}
        categories[cat]["count"] += 1
        categories[cat]["fmv"] += asset.get("fmv", 0) or 0
        categories[cat]["income"] += asset.get("annual_income", 0) or 0
    
    for cat_name, cat_data in categories.items():
        ws2.append([cat_name, cat_data["count"], cat_data["fmv"], cat_data["income"]])
    
    # Sheet 3: Summary
    ws3 = wb.create_sheet(_("sheet_summary", lang))
    ws3.append([_("summary_metric", lang), _("summary_value", lang)])
    style_header(ws3)
    
    total_fmv = sum(a.get("fmv", 0) or 0 for a in assets)
    total_income = sum(a.get("annual_income", 0) or 0 for a in assets)
    active_count = sum(1 for a in assets if a.get("status") == "Active")
    
    ws3.append([_("summary_total_assets", lang), len(assets)])
    ws3.append([_("summary_active_assets", lang), active_count])
    ws3.append([_("summary_total_fmv", lang), total_fmv])
    ws3.append([_("summary_total_income", lang), total_income])
    ws3.append([_("summary_total_categories", lang), len(categories)])
    ws3.append([_("last_updated", lang), datetime.now().isoformat()])
    
    # Sheet 4: Access
    ws4 = wb.create_sheet(_("sheet_access", lang))
    ws4.append([_("col_id", lang), _("col_asset_name", lang), _("col_institution", lang), _("col_login_url", lang),
                _("col_login_username", lang), _("col_login_password", lang), _("col_two_factor", lang)])
    style_header(ws4)
    
    for asset in assets:
        if asset.get("login_url") or asset.get("login_username"):
            ws4.append([
                asset.get("id", ""),
                asset.get("asset_name", ""),
                asset.get("institution", ""),
                asset.get("login_url", ""),
                asset.get("login_username", ""),
                asset.get("login_password", ""),
                asset.get("two_factor", ""),
            ])
    
    # Sheet 5: Financial Summary
    ws5 = wb.create_sheet(_("sheet_financial", lang))
    ws5.append([_("col_id", lang), _("col_asset_name", lang), _("col_currency", lang), _("col_acb", lang),
                _("col_fmv", lang), _("col_unrealized_gain", lang), _("col_unrealized_gain_pct", lang),
                _("col_annual_income", lang)])
    style_header(ws5)
    
    for asset in assets:
        acb = asset.get("acb", 0) or 0
        fmv = asset.get("fmv", 0) or 0
        gain = fmv - acb if acb > 0 else 0
        gain_pct = (gain / acb) if acb > 0 else 0
        
        ws5.append([
            asset.get("id", ""),
            asset.get("asset_name", ""),
            asset.get("currency", "CAD"),
            acb,
            fmv,
            gain,
            gain_pct,
            asset.get("annual_income", 0),
        ])
    
    # Sheet 6: Insurance
    ws6 = wb.create_sheet(_("sheet_insurance", lang))
    ws6.append([_("col_id", lang), _("col_asset_name", lang), _("col_insurance_coverage", lang), _("col_insurance_provider", lang),
                _("col_insurance_policy", lang), _("col_insurance_premium", lang), _("col_insured_value", lang)])
    style_header(ws6)
    
    for asset in assets:
        if asset.get("insurance_coverage") or asset.get("insurance_provider"):
            ws6.append([
                asset.get("id", ""),
                asset.get("asset_name", ""),
                asset.get("insurance_coverage", 0),
                asset.get("insurance_provider", ""),
                asset.get("insurance_policy", ""),
                asset.get("insurance_premium", 0),
                asset.get("insured_value", 0),
            ])
    
    # Sheet 7: Beneficiaries
    ws7 = wb.create_sheet(_("sheet_beneficiaries", lang))
    ws7.append([_("col_id", lang), _("col_asset_name", lang), _("col_owner", lang), _("col_primary_beneficiary", lang),
                _("col_beneficiary_pct_primary", lang), _("col_contingent_beneficiary", lang),
                _("col_beneficiary_pct_contingent", lang), _("col_pod", lang), _("col_tod", lang), _("col_poa", lang)])
    style_header(ws7)
    
    for asset in assets:
        if asset.get("primary_beneficiary") or asset.get("beneficiary"):
            ws7.append([
                asset.get("id", ""),
                asset.get("asset_name", ""),
                asset.get("owner", ""),
                asset.get("primary_beneficiary", "") or asset.get("beneficiary", ""),
                asset.get("beneficiary_pct_primary", 0),
                asset.get("contingent_beneficiary", ""),
                asset.get("beneficiary_pct_contingent", 0),
                asset.get("pod", ""),
                asset.get("tod", ""),
                asset.get("poa", ""),
            ])
    
    # Sheet 8: Estate
    ws8 = wb.create_sheet(_("sheet_estate", lang))
    ws8.append([_("col_id", lang), _("col_asset_name", lang), _("col_probate_excluded", lang), _("col_will_clause", lang),
                _("col_trust_clause", lang), _("col_estate_duty", lang), _("col_capital_gains_exemption", lang),
                _("col_succession_plan", lang), _("col_power_of_appointment", lang)])
    style_header(ws8)
    
    for asset in assets:
        if any([asset.get("probate_excluded"), asset.get("will_clause"), 
                asset.get("trust_clause"), asset.get("succession_plan")]):
            ws8.append([
                asset.get("id", ""),
                asset.get("asset_name", ""),
                asset.get("probate_excluded", ""),
                asset.get("will_clause", ""),
                asset.get("trust_clause", ""),
                asset.get("estate_duty", ""),
                asset.get("capital_gains_exemption", ""),
                asset.get("succession_plan", ""),
                asset.get("power_of_appointment", ""),
            ])
    
    # Save workbook
    wb.save(output_path)
    print(f"[OK] Excel generated: {output_path}")

# =============================================================================
# HTML DASHBOARD GENERATOR
# =============================================================================

def generate_html(assets: list[dict], output_path: str, lang: str = "en",
                  tier: str = "free", license_token: str = "", buyer: str = "",
                  key_version: int = 1, license_secret: str = ""):
    """Generate self-contained HTML dashboard with Modern Financial Institution design."""

    # Translate field definitions (labels and group names) for the modal/detail view
    translated_field_defs = translate_field_definitions(FIELD_DEFINITIONS, lang)

    # Build category metadata using translated category names while keeping
    # original English lookup for icon/color from ASSET_CATEGORIES.
    categories = {}
    reverse_cat_map = {v: k for k, v in CATEGORY_TRANSLATIONS.get(lang, {}).items()}
    for asset in assets:
        cat = asset["category"]
        source_cat = reverse_cat_map.get(cat, cat)
        if cat not in categories:
            categories[cat] = {
                "count": 0,
                "icon": ASSET_CATEGORIES.get(source_cat, {}).get("icon", "📋"),
                "color": ASSET_CATEGORIES.get(source_cat, {}).get("color", "#6c5ce7"),
            }
        categories[cat]["count"] += 1

    # Serialize data to JSON for embedding.
    inventory = {
        "format": "asset-inventory",
        "version": 2,
        "schema_version": 1,
        "tier": tier,
        "key_version": key_version,
        "generated": datetime.now().strftime("%Y-%m-%d"),
        "assets": assets,
    }
    inventory_json = json.dumps(inventory, indent=2, default=str)
    categories_json = json.dumps(categories, indent=2)
    fields_json = json.dumps(translated_field_defs, indent=2)

    template_path = Path(__file__).parent.parent / "templates" / "dashboard.html"
    template = template_path.read_text(encoding="utf-8")

    # Build-time tier stripping: lower tiers physically remove higher-tier code.
    # Markers in the template:
    #   <!--__TIER_GE:family--> ... <!--__/TIER_GE:family-->
    #   <!--__TIER_GE:planning--> ... <!--__/TIER_GE:planning-->
    tier_rank = {"free": 0, "family": 1, "planning": 2, "advisor": 3}
    cur = tier_rank.get(tier, 0)
    html = template
    for name in ("family", "planning", "advisor"):
        if cur < tier_rank[name]:
            open_m, close_m = f"<!--__TIER_GE:{name}-->", f"<!--__/TIER_GE:{name}-->"
            while True:
                s = html.find(open_m)
                if s < 0:
                    break
                e = html.find(close_m, s)
                if e < 0:
                    break
                e += len(close_m)
                html = html[:s] + html[e:]
    # Strip any leftover marker comment tokens.
    html = re.sub(r"<!--__/?TIER_GE:[a-z]+-->", "", html)

    # Substitute UI translation placeholders first so they cannot be clobbered
    # by JSON content.
    for key, value in UI_TRANSLATIONS.get(lang, UI_TRANSLATIONS["en"]).items():
        html = html.replace(f"{{{{TR_{key}}}}}", str(value))

    html = (html
        .replace("{{INVENTORY_JSON}}", inventory_json)
        .replace("{{TIER}}", tier)
        .replace("{{LICENSE_TOKEN}}", license_token)
        .replace("{{LICENSE_JSON}}", license_token)
        .replace("{{LICENSE_SECRET}}", license_secret)
        .replace("{{BUYER}}", buyer)
        .replace("{{KEY_VERSION}}", str(key_version))
        .replace("{{CATEGORIES_JSON}}", categories_json)
        .replace("{{FIELDS_JSON}}", fields_json))

    # Write file
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"[OK] HTML generated: {output_path}")

# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Asset Inventory Generator for Canadian Families in Ontario"
    )
    parser.add_argument(
        "--output", "-o",
        choices=["md", "excel", "html", "all"],
        default="all",
        help="Output format (default: all)"
    )
    parser.add_argument(
        "--output-dir", "-d",
        default="./output",
        help="Output directory (default: ./output)"
    )
    parser.add_argument(
        "--owner",
        help="Filter by owner name"
    )
    parser.add_argument(
        "--category",
        help="Filter by category"
    )
    parser.add_argument(
        "--status",
        choices=["active", "dormant", "all"],
        default="all",
        help="Filter by status"
    )
    parser.add_argument(
        "--lang", "-l",
        choices=["en", "zh"],
        default="en",
        help="Output language (default: en)"
    )
    parser.add_argument(
        "--demo",
        action="store_true",
        help="Overlay a realistic Ontario family demo fixture (James & Mei Chen, "
             "house + mortgage, TFSAs/RRSPs/RESP, dormant crypto wallet)"
    )
    parser.add_argument(
        "--tier",
        choices=["free", "family", "planning", "advisor"],
        default="free",
        help="Edition tier (default: free). Free/Family = HTML only; Planning = HTML+MD+Excel"
    )
    parser.add_argument(
        "--license",
        default="",
        help="Signed license token for paid tiers (embedded in the file)"
    )
    parser.add_argument(
        "--buyer",
        default="",
        help='Purchaser identity for watermark, e.g. "Name <email>" (paid tiers)'
    )
    parser.add_argument(
        "--key-version",
        type=int,
        default=1,
        help="Data-block decryption key chain version (default: 1)"
    )
    parser.add_argument(
        "--license-secret",
        default="",
        help="Secret used to sign the embedded license (paid tiers; keep private)"
    )
    parser.add_argument(
        "--expires",
        default="",
        help="Optional license expiry (ISO date) for update packs; empty = perpetual"
    )
    
    args = parser.parse_args()
    
    # Create output directory
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Generate assets
    print("\n" + "="*60)
    print("ASSET INVENTORY GENERATOR")
    print("Canadian Family in Ontario")
    print("="*60 + "\n")
    
    print(f"Generating assets (tier: {args.tier})...")
    assets = generate_all_assets(tier=args.tier)
    tier_cat_count = len(TIER_CATEGORIES.get(args.tier, TIER_CATEGORIES["planning"]))
    print(f"Generated {len(assets)} assets across {tier_cat_count} categories")
    
    # Demo fixture: realistic family data overlays the blank template
    if args.demo:
        assets = populate_demo_data(assets)
        print(f"[DEMO] Populated demo fixture ({sum(1 for a in assets if a.get('source') == 'demo')} assets with sample data)")
    
    # Apply filters before translation so filters match source data
    if args.owner:
        assets = [a for a in assets if args.owner.lower() in (a.get("owner", "") or "").lower()]
        print(f"Filtered to {len(assets)} assets (owner: {args.owner})")
    
    if args.category:
        assets = [a for a in assets if args.category.lower() in a.get("category", "").lower()]
        print(f"Filtered to {len(assets)} assets (category: {args.category})")
    
    if args.status != "all":
        status_map = {"active": "Active", "dormant": "Dormant"}
        target_status = status_map.get(args.status, args.status)
        assets = [a for a in assets if a.get("status") == target_status]
        print(f"Filtered to {len(assets)} assets (status: {target_status})")
    
    # Translate assets and metadata for non-English output
    lang = args.lang
    translated_assets = translate_assets(assets, lang)
    
    # Build language suffix for output filenames
    lang_suffix = "-zh" if lang == "zh" else ""
    tier_suffix = f"-{args.tier}" if args.tier != "free" else ""
    
    # Export (Markdown/Excel) is Planning-tier only per versioning_plan.
    export_enabled = args.tier in ("planning", "advisor")
    
    # Generate outputs
    print("\nGenerating outputs...")
    
    if export_enabled and args.output in ["md", "all"]:
        md_path = output_dir / f"asset-inventory{lang_suffix}{tier_suffix}.md"
        generate_markdown(translated_assets, str(md_path), lang=lang)
    
    if export_enabled and args.output in ["excel", "all"]:
        excel_path = output_dir / f"asset-inventory{lang_suffix}{tier_suffix}.xlsx"
        generate_excel(translated_assets, str(excel_path), lang=lang)
    
    if args.output in ["html", "all"]:
        html_path = output_dir / f"asset-inventory-dashboard{lang_suffix}{tier_suffix}.html"
        # Build the signed license for paid tiers.
        license_secret = args.license_secret or "dev-secret"
        license_token = args.license
        if args.tier in ("family", "planning", "advisor") and not license_token:
            license_payload = {
                "tier": args.tier,
                "order_id": "",
                "buyer": args.buyer or "",
                "issued": datetime.now().strftime("%Y-%m-%d"),
            }
            if args.expires:
                license_payload["expires"] = args.expires
            license_token = sign_license(license_payload, license_secret)
            print(f"[LICENSE] Signed {args.tier} license embedded")
        generate_html(translated_assets, str(html_path), lang=lang, tier=args.tier,
                      license_token=license_token, buyer=args.buyer,
                      key_version=args.key_version, license_secret=license_secret)
    
    print("\n" + "="*60)
    print("GENERATION COMPLETE")
    print("="*60)
    print(f"Total Assets: {len(assets)}")
    print(f"Categories: {tier_cat_count}")
    print(f"Fields per asset: {len(FIELD_DEFINITIONS)}")
    print(f"\nOutputs saved to: {output_dir.absolute()}")
    print("="*60 + "\n")

if __name__ == "__main__":
    main()
