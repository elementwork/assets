#!/usr/bin/env python3
"""
End-to-end visual test for the Asset Inventory Generator.

Pipeline under test:
    1. Regenerate all outputs (en + zh): Markdown, Excel, self-contained HTML dashboard
    2. Static validation of the generated artifacts (no leftover template placeholders,
       Excel workbook structure, Markdown content)
    3. Headless-Chromium interaction on the generated dashboards (en + zh):
       stats, search, filters, themes, templates, layouts, table sorting, edit modal,
       validation, undo/redo, duplicate, delete, file lock (whole-file encryption),
       charts, exports, import, print
    4. Screenshots captured to tests/screenshots/ for visual review

Exit code 0 = all checks pass, 1 = one or more failures.
"""

import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path

from playwright.sync_api import sync_playwright

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "output"
SHOTS = ROOT / "tests" / "screenshots"
TMP = Path(tempfile.mkdtemp(prefix="e2e-tiers-"))

# Main E2E runs against the Planning tier (full feature set: table/charts/export/
# print + MD/Excel artifacts). Tier-gating differences are asserted separately
# in tier_gating_check().
EN_HTML = OUT / "asset-inventory-dashboard-planning.html"
ZH_HTML = OUT / "asset-inventory-dashboard-zh-planning.html"
EN_MD = OUT / "asset-inventory-planning.md"
ZH_MD = OUT / "asset-inventory-zh-planning.md"
EN_XLSX = OUT / "asset-inventory-planning.xlsx"
ZH_XLSX = OUT / "asset-inventory-zh-planning.xlsx"

PASSED = []
FAILED = []


def check(name, cond, detail=""):
    if cond:
        PASSED.append(name)
        print(f"  PASS  {name}")
    else:
        FAILED.append(name)
        print(f"  FAIL  {name}  {detail}")
    return cond


# =============================================================================
# STEP 1 — Regenerate outputs
# =============================================================================

def run_generation():
    print("\n== Step 1: regenerate outputs (en + zh, planning tier) ==")
    for lang in ("en", "zh"):
        cmd = [sys.executable, str(ROOT / "src" / "generate_asset_inventory.py"),
               "--tier", "planning"]
        if lang == "zh":
            cmd += ["--lang", "zh"]
        r = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, timeout=300)
        check(
            f"generator ({lang}) exits 0",
            r.returncode == 0,
            f"exit={r.returncode}\nstdout={r.stdout[-800:]}\nstderr={r.stderr[-800:]}",
        )
    for p in (EN_MD, ZH_MD, EN_XLSX, ZH_XLSX, EN_HTML, ZH_HTML):
        check(f"output exists: {p.name}", p.exists() and p.stat().st_size > 0)


# =============================================================================
# STEP 1b — Tier gating (Free/Family/Planning)
# =============================================================================

def tier_gating_check():
    print("\n== Step 1b: tier gating ==")
    cases = {
        "free":    {"assets": 256, "tpl": 1, "edition_accent": "#2563EB", "table": False, "timeline": False,
                    "charts": False, "audit": False, "review": False, "print": False, "export": False},
        "family":  {"assets": 324, "tpl": 5, "edition_accent": "#A66E14", "table": True, "timeline": True,
                    "charts": True, "audit": False, "review": False, "print": True, "export": False},
        "planning": {"assets": 517, "tpl": 5, "edition_accent": "#087E79", "table": True, "timeline": True,
                     "charts": True, "audit": True, "review": True, "print": True, "export": True},
    }
    for tier, cfg in cases.items():
        cmd = [sys.executable, str(ROOT / "src" / "generate_asset_inventory.py"),
               "--tier", tier, "-o", "html", "-d", str(TMP)]
        r = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, timeout=300)
        check(f"tier {tier}: generator exits 0", r.returncode == 0, r.stderr[-400:])
        fname = "asset-inventory-dashboard.html" if tier == "free" \
            else f"asset-inventory-dashboard-{tier}.html"
        html = (TMP / fname).read_text(encoding="utf-8")
        check(f"tier {tier}: asset count embedded", f'"id": "A-' in html)
        # Lower tiers must NOT contain higher-tier code (split build).
        if tier == "free":
            check("free: no higher-tier export code",
                  all(t not in html for t in ["function exportMarkdown",
                                               "function exportExcel", "function exportJSON"]))
            check("free: no print/charts/audit code",
                  all(t not in html for t in ["function renderPrintView",
                                               "function renderCharts", "function renderAudit",
                                               "function renderTable", "function renderTimeline"]))
        if tier == "family":
            check("family: no planning-only export code",
                  all(t not in html for t in ["function exportMarkdown",
                                               "function exportExcel", "function exportJSON"]))
            check("family: charts/timeline/print present, audit stripped",
                  all(t in html for t in ["function renderCharts",
                                          "function renderTimeline", "function renderPrintView"])
                  and "function renderAudit" not in html)
        with sync_playwright() as p:
            browser = p.chromium.launch()
            ctx = browser.new_context()
            page = ctx.new_page()
            errs = []
            page.on("pageerror", lambda e: errs.append(str(e)))
            page.goto((TMP / fname).as_uri(), wait_until="load")
            page.wait_for_timeout(600)
            check(f"tier {tier}: totalAssets={cfg['assets']}",
                  page.text_content("#totalAssets") == str(cfg["assets"]),
                  page.text_content("#totalAssets"))
            check(f"tier {tier}: template count={cfg['tpl']}",
                  page.locator("#templateSelect option").count() == cfg["tpl"])
            check(f"tier {tier}: edition visual identity",
                  page.get_attribute("html", "data-edition") == tier
                  and page.evaluate("getComputedStyle(document.documentElement).getPropertyValue('--edition-accent').trim()") == cfg["edition_accent"],
                  page.evaluate("JSON.stringify({edition: document.documentElement.dataset.edition, accent: getComputedStyle(document.documentElement).getPropertyValue('--edition-accent').trim()})"))
            page.screenshot(path=str(SHOTS / f"tier-{tier}-dashboard.png"))
            for feat, on in cfg.items():
                if feat in ("assets", "tpl", "edition_accent"):
                    continue
                selector = {"print": "#printBtn", "export": "#exportMD"}.get(feat)
                if selector:
                    available = page.locator(selector).count() > 0
                    check(f"tier {tier}: {feat} available={on}", available == on, f"got {available}")
                else:
                    available = page.locator(f'.header-layout-btn[data-layout="{feat}"]').count() > 0
                    check(f"tier {tier}: layout {feat} available={on}", available == on, f"got {available}")
            check(f"tier {tier}: edition badge", bool(page.text_content("#editionBadge")))
            teaser_count = page.locator('.upgrade-preview.visible').count()
            expected_teasers = 2 if tier == 'free' else (1 if tier == 'family' else 0)
            check(f"tier {tier}: upgrade preview count={expected_teasers}", teaser_count == expected_teasers, f"got {teaser_count}")
            check(f"tier {tier}: no JS errors", not errs, f"{errs[:3]}")
            ctx.close()
            browser.close()


def license_downgrade_ui_check():
    print("\n== Step 1c: invalid-license effective-tier UI ==")
    path = TMP / "asset-inventory-dashboard-planning.html"
    if not path.exists():
        check("invalid license fixture exists", False, str(path))
        return
    html = path.read_text(encoding="utf-8")
    tampered = re.sub(r"const LICENSE_JSON = '[^']*';",
                      "const LICENSE_JSON = 'invalid.invalid';", html, count=1)
    invalid = TMP / "asset-inventory-dashboard-planning-invalid-license.html"
    invalid.write_text(tampered, encoding="utf-8")
    with sync_playwright() as p:
        browser = p.chromium.launch()
        ctx = browser.new_context()
        ctx.add_init_script("localStorage.setItem('assetLayout', 'audit')")
        page = ctx.new_page()
        errors = []
        page.on("pageerror", lambda exc: errors.append(str(exc)))
        page.goto(invalid.as_uri(), wait_until="load")
        page.wait_for_timeout(800)
        check("invalid license: effective edition is Free",
              page.get_attribute("html", "data-edition") == "free",
              page.get_attribute("html", "data-edition"))
        check("invalid license: Free badge shown", "Free" in page.text_content("#editionBadge"), page.text_content("#editionBadge"))
        check("invalid license: stale Audit layout normalized",
              page.evaluate("currentLayout") == "dashboard" and page.locator("#dashboardView.active").count() == 1,
              page.evaluate("currentLayout"))
        check("invalid license: print hidden", page.locator("#printBtn").count() == 1 and not page.locator("#printBtn").is_visible())
        check("invalid license: exports hidden", page.locator("#exportMD").count() == 1 and not page.locator("#exportMD").is_visible())
        check("invalid license: Family tools hidden", page.locator('[data-min-tier="family"]:visible').count() == 0)
        check("invalid license: template reduced to one", page.locator("#templateSelect option").count() == 1)
        check("invalid license: no JS errors", not errors, str(errors[:3]))
        ctx.close()
        browser.close()


# =============================================================================
# STEP 2 — Static validation of generated artifacts
# =============================================================================

def static_checks():
    print("\n== Step 2: static artifact checks ==")

    # Tier markers in the template must be balanced and properly nested.
    # The generator now fails hard on imbalance; this check guards the source
    # template deterministically so drift fails CI with a clear message.
    tpl_lines = (ROOT / "templates" / "dashboard.html").read_text(encoding="utf-8").splitlines()
    stack = []
    marker_ok = True
    for i, ln in enumerate(tpl_lines, 1):
        m = re.search(r"<!--__(/)?TIER_GE:([a-z]+)-->", ln)
        if not m:
            continue
        is_close, name = bool(m.group(1)), m.group(2)
        if is_close:
            if not stack or stack[-1][0] != name:
                marker_ok = False
                print(f"  marker mismatch at line {i}: {ln.strip()}")
                break
            stack.pop()
        else:
            stack.append((name, i))
    check("template tier markers balanced & nested", marker_ok and not stack,
          f"unclosed={stack}")

    en_html = EN_HTML.read_text(encoding="utf-8")
    zh_html = ZH_HTML.read_text(encoding="utf-8")
    check("inventory identity embedded", '"inventory_id": "INV-' in en_html)
    check("direct-save API path compiled", "showSaveFilePicker" in en_html and "directSaveCurrentFile" in en_html)
    check("persistent file-handle binding compiled",
          all(t in en_html for t in ["indexedDB.open", "persistSaveFileHandle", "loadPersistedSaveFileHandle", "restoreSaveFileHandle", "ensureWritePermission"]))
    check("planning annual-review layout compiled", 'data-layout="review"' in en_html and "renderAnnualReview" in en_html)
    check("tier-aware feature menu compiled", all(t in en_html for t in ["featureMenuToggle", "featureMenuClose", "editionBadge", "upgrade-preview", "scope-filter-btn"]))
    check("effective-tier normalization compiled", all(t in en_html for t in ["effectiveTier", "normalizeLayoutForTier", "cloneTierConfig", "downgradeToFree"]))
    check("simplified chrome compiled", all(t in en_html for t in ["logo-menu-toggle", "statAddAssetBtn", "filter-status-toggles", "header-view-switcher"]))
    check("compound filter controls compiled", all(t in en_html for t in ["role=\"group\"", "quick-scope-filters", "filter-status-toggles", "header-primary-action"]))
    check("standalone workspace/layout switcher removed", '<div class="workspace-bar"' not in en_html and '<div class="layout-switcher"' not in en_html)
    check("professional terminology retained", all(t in en_html for t in ['aria-label="Audit"', 'aria-label="Annual Review"', "Export MD", "Export JSON"]))
    check("handoff schema fields compiled", all(k in en_html for k in ["emergency_priority", "incapacity_access", "death_access", "last_access_test"]))
    check("zh continuity strings compiled", "紧急访问指南" in zh_html and "家庭年度复核" in zh_html)

    # Markdown
    en_md = EN_MD.read_text(encoding="utf-8")
    zh_md = ZH_MD.read_text(encoding="utf-8")
    check("en md: title header", "# Asset Inventory" in en_md)
    check("en md: 517 assets reported", "Total Assets:** 517" in en_md)
    check("en md: category sections", en_md.count("\n## ") >= 32)
    check("zh md: Chinese title", "资产清单" in zh_md)
    check("zh md: 517 assets reported", "资产总数:** 517" in zh_md)

    # Excel
    try:
        import openpyxl
    except ImportError:
        check("openpyxl importable", False, "pip install openpyxl")
        return
    wb = openpyxl.load_workbook(EN_XLSX)
    en_sheets = [ws.title for ws in wb.worksheets]
    check(
        "en xlsx: expected sheets",
        en_sheets == ["All Assets", "By Category", "Summary", "Access",
                      "Financial Summary", "Insurance", "Beneficiaries", "Estate"],
        f"got {en_sheets}",
    )
    check("en xlsx: All Assets has 518 rows", wb["All Assets"].max_row == 518)
    wb_zh = openpyxl.load_workbook(ZH_XLSX)
    zh_sheets = [ws.title for ws in wb_zh.worksheets]
    check(
        "zh xlsx: expected sheets",
        zh_sheets == ["全部资产", "按类别", "摘要", "访问信息",
                      "财务摘要", "保险", "受益人", "遗产"],
        f"got {zh_sheets}",
    )
    check("zh xlsx: 全部资产 has 518 rows", wb_zh["全部资产"].max_row == 518)

    # HTML: no leftover placeholders, data blocks present
    for label, path in (("en", EN_HTML), ("zh", ZH_HTML)):
        html = path.read_text(encoding="utf-8")
        leftovers = re.findall(r"\{\{(?:TR_|INVENTORY_JSON|CATEGORIES_JSON|FIELDS_JSON)", html)
        check(f"{label} html: no leftover placeholders", not leftovers, f"found {leftovers[:5]}")
        check(f"{label} html: INVENTORY_DATA embedded", "const INVENTORY_DATA = {" in html)
        check(f"{label} html: CATEGORIES_DATA embedded", "const CATEGORIES_DATA =" in html)
        check(f"{label} html: FIELDS_DATA embedded", "const FIELDS_DATA =" in html)
        check(f"{label} html: 517 assets in JSON", html.count('"id": "A-') == 517)
        check(f"{label} html: file-size sanity (>200KB)", len(html) > 200_000)


# =============================================================================
# STEP 2b — Demo fixture (1.7)
# =============================================================================

def demo_fixture_check():
    print("\n== Step 2b: --demo fixture (1.7, planning tier) ==")
    with tempfile.TemporaryDirectory() as tmp:
        cmd = [sys.executable, str(ROOT / "src" / "generate_asset_inventory.py"),
               "--tier", "planning", "--demo", "-d", tmp]
        r = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, timeout=300)
        check("demo: generator --demo exits 0", r.returncode == 0, r.stderr[-500:])
        html = (Path(tmp) / "asset-inventory-dashboard-planning.html").read_text(encoding="utf-8")
        i = html.find('"assets": [')
        if i < 0:
            check("demo: INVENTORY_DATA assets found", False, "no assets array")
            return
        depth, pos = 0, i
        while pos < len(html):
            if html[pos] == '[':
                depth += 1
            elif html[pos] == ']':
                depth -= 1
                if depth == 0:
                    break
            pos += 1
        # Include the opening '[' so the slice is a balanced array.
        data = json.loads(html[i + len('"assets": '):pos + 1])
        demo = [a for a in data if a.get("source") == "demo"]
        check("demo: 27 populated assets", len(demo) == 27, f"n={len(demo)}")
        total_fmv = sum(float(a.get("fmv") or 0) for a in demo)
        check("demo: FMV populated", total_fmv > 2_000_000, f"${total_fmv:,.0f}")
        owners = {a.get("owner") for a in demo if a.get("owner")}
        check("demo: James & Mei Chen owners",
              {"James Chen", "Mei Chen"} <= owners, str(owners))
        dormant = [a for a in demo if a.get("status") == "Dormant"]
        check("demo: dormant crypto wallet",
              any("Bitcoin" in (a.get("asset_name") or "") for a in dormant))
        check("demo: markdown notes in fixture",
              any("**" in (a.get("notes") or "") for a in demo))


# =============================================================================
# STEP 3 — Browser E2E
# =============================================================================

def responsive_ui_check():
    print("\n== Step 2c: responsive centered header + single-line filters ==")
    with sync_playwright() as p:
        browser = p.chromium.launch()
        for width, height in ((390, 844), (768, 900), (1440, 1000)):
            ctx = browser.new_context(viewport={"width": width, "height": height})
            page = ctx.new_page()
            errors = []
            page.on("pageerror", lambda exc: errors.append(str(exc)))
            page.goto(EN_HTML.as_uri(), wait_until="load")
            page.wait_for_timeout(500)
            check(f"responsive {width}: header Add removed",
                  page.locator(".tier-header #addAssetBtn").count() == 0)
            check(f"responsive {width}: right side is Last Updated + Save",
                  page.locator(".tier-header #lastUpdated").count() == 1
                  and page.locator(".tier-header #saveHTML").count() == 1)
            check(f"responsive {width}: centered SVG view switcher present",
                  page.locator("#headerViewSwitcher .header-layout-btn svg").count() == 9)
            box = page.locator("#headerViewSwitcher").bounding_box()
            check(f"responsive {width}: view switcher centered in viewport",
                  bool(box and abs((box['x'] + box['width'] / 2) - width / 2) <= 3), str(box))
            check(f"responsive {width}: all filters use one nowrap bar",
                  page.locator(".filter-bar").evaluate("el => getComputedStyle(el).flexWrap") == "nowrap")
            status_boxes = page.locator(".filter-bar .status-toggle-btn").evaluate_all(
                "els => els.map(el => { const r=el.getBoundingClientRect(); return [r.top,r.bottom]; })")
            check(f"responsive {width}: status buttons do not wrap",
                  len(status_boxes) == 4 and max(b[0] for b in status_boxes) - min(b[0] for b in status_boxes) <= 2,
                  str(status_boxes))
            check(f"responsive {width}: search/category/owner/status all in filter bar",
                  page.locator(".filter-bar #searchInput").count() == 1
                  and page.locator(".filter-bar #categoryFilter").count() == 1
                  and page.locator(".filter-bar #ownerFilter").count() == 1
                  and page.locator(".filter-bar .status-toggle-btn").count() == 4)
            check(f"responsive {width}: filter result count absent",
                  page.locator(".filter-bar .filter-result-count").count() == 0)
            asset_card = page.locator(".stat-card-assets")
            left_box = asset_card.locator(".assets-stat-left").bounding_box()
            right_box = asset_card.locator(".assets-stat-right").bounding_box()
            check(f"responsive {width}: Assets card stays two-column on one row",
                  bool(left_box and right_box and abs((left_box['y'] + left_box['height']/2) - (right_box['y'] + right_box['height']/2)) <= 18
                       and left_box['x'] < right_box['x']),
                  f"left={left_box} right={right_box}")
            check(f"responsive {width}: Assets card right column order",
                  asset_card.locator(".assets-stat-right > *").count() == 2
                  and asset_card.locator(".assets-stat-right > *").nth(0).get_attribute("id") == "statAddAssetBtn"
                  and asset_card.locator(".assets-stat-right > *").nth(1).get_attribute("class").startswith("stat-subline"))
            page.click("#featureMenuToggle")
            check(f"responsive {width}: logo menu opens", page.locator("#featureMenu.open").count() == 1)
            check(f"responsive {width}: Views also present in menu",
                  page.locator("#featureMenu .menu-layout-item").count() == 9
                  and page.locator("#featureMenu .menu-layout-item svg").count() == 9)
            page.locator('#featureMenu .menu-layout-item[data-layout="kanban"]').click()
            page.wait_for_timeout(120)
            check(f"responsive {width}: menu View syncs header active state",
                  page.locator('#headerViewSwitcher .header-layout-btn[data-layout="kanban"].active').count() == 1)
            page.click("#featureMenuToggle")
            page.locator('#featureMenu .menu-layout-item[data-layout="dashboard"]').click()
            page.wait_for_timeout(120)
            check(f"responsive {width}: no JS errors", not errors, str(errors[:3]))
            ctx.close()
        browser.close()


def browser_e2e():
    print("\n== Step 3: browser E2E (en + zh) ==")
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        ok = e2e_en(browser)
        ok = e2e_zh(browser) and ok
        browser.close()
        return ok


def _new_page(browser, html_path, viewport=(1440, 1000)):
    context = browser.new_context(viewport={"width": viewport[0], "height": viewport[1]},
                                  accept_downloads=True)
    page = context.new_page()
    js_errors = []
    page.on("pageerror", lambda exc: js_errors.append(str(exc)))
    page.on("console", lambda msg: js_errors.append(msg.text) if msg.type == "error" else None)
    page.goto(html_path.as_uri(), wait_until="load")
    return context, page, js_errors


def _open_feature_menu(page):
    menu = page.locator("#featureMenu")
    if not menu.evaluate("el => el.classList.contains('open')"):
        page.click("#featureMenuToggle")
        page.wait_for_timeout(100)


def _menu_click(page, selector):
    _open_feature_menu(page)
    page.click(selector)
    page.wait_for_timeout(120)


def _layout(page, name):
    primary = page.locator(f'.layout-btn[data-layout="{name}"]:visible')
    if primary.count():
        primary.first.click()
    else:
        _open_feature_menu(page)
        page.locator(f'.menu-layout-item[data-layout="{name}"]:visible').first.click()
    page.wait_for_timeout(150)


def _open_modal(page, asset_id):
    page.click(f'[data-asset-id="{asset_id}"]')
    page.wait_for_selector("#modalOverlay.active", timeout=5000)


def _save_modal(page):
    page.click("#modalSave")
    page.wait_for_selector("#modalOverlay:not(.active)", timeout=8000)


def e2e_en(browser):
    print("  -- en dashboard --")
    context, page, js_errors = _new_page(browser, EN_HTML)
    ok = True

    # Core rendering
    check("en: page title", "Asset Inventory" in page.title(), page.title())
    check("en: stat totalAssets = 517", page.text_content("#totalAssets") == "517",
          page.text_content("#totalAssets"))
    check("en: stat totalCategories = 32", page.text_content("#totalCategories") == "32",
          page.text_content("#totalCategories"))
    check("en: professional edition badge", "Professional" in page.text_content("#editionBadge"))
    check("en: blank catalog has zero assets with value", page.text_content("#withValueCount") == "0", page.text_content("#withValueCount"))
    check("en: Asset Types label", "Asset Types" in page.locator(".stat-card").first.text_content())
    check("en: blank catalog starts with zero recorded Assets", page.text_content("#assetCount") == "0", page.text_content("#assetCount"))
    check("en: stat card order is Asset Types / Categories / Assets", [x.strip() for x in page.locator(".primary-stats .stat-label").all_text_contents()[:3]] == ["Asset Types", "Categories", "Assets"])
    check("en: filter result count removed", page.locator(".filter-result-count").count() == 0 and page.locator("#showingCount").count() == 0)
    check("en: quick scope filters render", page.locator(".scope-filter-btn").count() == 2)
    check("en: USD FMV counts as With Value", page.evaluate("hasFinancialValue({fmv_usd: 100})"))
    check("en: dashboard view renders 517 items",
          page.locator("#dashboardView .asset-item").count() == 517)

    # Search (input is debounced 300ms in the template)
    page.fill("#searchInput", "Chequing")
    page.wait_for_timeout(600)
    n = page.locator("#dashboardView .asset-item").count()
    check("en: search 'Chequing' narrows results", 0 < n < 517, f"n={n}")
    page.fill("#searchInput", "")
    page.wait_for_timeout(600)
    check("en: clearing search restores 517",
          page.locator("#dashboardView .asset-item").count() == 517)

    # Category filter
    first_cat = page.locator("#categoryFilter option").nth(1).text_content()
    expected = page.evaluate(
        "assets.filter(a => a.category === document.querySelector('#categoryFilter').value).length"
    ) if False else None
    page.select_option("#categoryFilter", index=1)
    page.wait_for_timeout(150)
    actual = page.locator("#dashboardView .asset-item").count()
    # Compare with the JS-side filtered length for correctness, not just a drop.
    expected = page.evaluate(
        "getFilteredAssets().length"
    )
    check(f"en: category filter '{first_cat}' matches", actual == expected,
          f"dom={actual} js={expected}")
    page.select_option("#categoryFilter", index=0)
    page.wait_for_timeout(150)

    # Feature menu close affordance + focus return
    _open_feature_menu(page)
    page.locator("#featureMenuClose").focus()
    page.click("#featureMenuClose")
    check("en: feature menu close button works", page.locator("#featureMenu.open").count() == 0)
    check("en: feature menu returns focus to toggle", page.evaluate("document.activeElement && document.activeElement.id") == "featureMenuToggle")

    # Theme toggle
    theme_before = page.evaluate("document.documentElement.getAttribute('data-theme')")
    _menu_click(page, "#themeToggle")
    theme_after = page.evaluate("document.documentElement.getAttribute('data-theme')")
    check("en: theme toggle flips light/dark", theme_before != theme_after,
          f"{theme_before} -> {theme_after}")
    _menu_click(page, "#themeToggle")  # back to light

    # Template select (5 templates) — now organized under Features > Appearance.
    _open_feature_menu(page)
    templates = page.locator("#templateSelect option").count()
    check("en: 5 template options", templates == 5, f"n={templates}")
    for i in range(1, 5):  # indexes 1..4 cover all 5 templates (0..4)
        page.select_option("#templateSelect", index=i)
        attr = page.evaluate("document.documentElement.getAttribute('data-template')")
        if i == 1:
            check("en: template select updates data-template", bool(attr), attr)
    page.screenshot(path=str(SHOTS / "en-dashboard.png"))

    # Layouts
    _layout(page, "table")
    check("en: table layout rows (50/page)", page.locator("#tableBody tr").count() == 50,
          page.locator("#tableBody tr").count())
    first_id_before = page.locator("#tableBody tr").first.text_content()
    page.click('th[data-sort="asset_name"]')  # sort ascending by name
    page.wait_for_timeout(150)
    page.click('th[data-sort="asset_name"]')  # sort descending
    page.wait_for_timeout(150)
    first_id_after = page.locator("#tableBody tr").first.text_content()
    check("en: table sort reorders rows", first_id_before != first_id_after,
          f"\nbefore={first_id_before[:60]!r}\nafter={first_id_after[:60]!r}")
    page.screenshot(path=str(SHOTS / "en-table.png"))

    _layout(page, "kanban")
    check("en: kanban cards", page.locator(".kanban-card").count() == 517)
    page.screenshot(path=str(SHOTS / "en-kanban.png"))

    _layout(page, "timeline")
    tl = page.locator(".timeline-item").count()
    check("en: timeline renders (items or empty state)", tl >= 0)
    check("en: timeline empty-state shown for undated template",
          page.locator(".empty-state").count() >= 1 or tl > 0)

    _layout(page, "detail")
    check("en: detail list items", page.locator("#detailAssetList .detail-asset-item").count() == 517)
    check("en: detail content grid", page.locator("#detailContent .detail-grid").count() == 1)

    _layout(page, "compact")
    check("en: compact view items", page.locator(".compact-asset").count() == 517)

    _layout(page, "audit")
    check("en: audit summary pills", page.locator(".audit-summary .audit-pill").count() == 3)
    n_audit = page.locator(".audit-item").count()
    check("en: audit flags all 517 template assets", n_audit == 517, f"n={n_audit}")
    page.screenshot(path=str(SHOTS / "en-audit.png"))

    # Charts: template has no FMV -> empty charts; after an FMV edit they populate.
    _layout(page, "charts")
    check("en: charts cards render", page.locator(".chart-card").count() == 4)
    check("en: charts empty without FMV data", page.locator(".chart-empty").count() >= 1)

    # Edit modal: schema-driven inputs
    _layout(page, "dashboard")
    _open_modal(page, "A-0001")
    fields = page.locator("#modalContent [data-field]").count()
    check("en: modal renders fields", fields > 50, f"n={fields}")
    check("en: currency input is type=number",
          page.locator('#modalContent [data-field="fmv"]').get_attribute("type") == "number")
    check("en: percent input is type=number",
          page.locator('#modalContent [data-field="yield_pct"]').get_attribute("type") == "number")
    date_field = page.locator('#modalContent [data-field="acquisition_date"]')
    check("en: date input is type=date",
          date_field.count() == 1 and date_field.get_attribute("type") == "date")
    check("en: select input rendered",
          page.locator('#modalContent select[data-field="status"]').count() == 1)

    # Validation: required asset_name blocks save
    page.fill('#modalContent [data-field="asset_name"]', "")
    page.click("#modalSave")
    page.wait_for_timeout(200)
    check("en: validation blocks save on missing name",
          page.locator("#modalOverlay.active").count() == 1 and
          page.locator("#modalContent .form-error").count() >= 1)
    page.fill('#modalContent [data-field="asset_name"]', "Test Chequing Account")

    # Edit + save: FMV so charts populate; name so we can verify persistence
    page.fill('#modalContent [data-field="fmv"]', "25000")
    _save_modal(page)
    check("en: unsaved-changes indicator appears",
          page.locator("#unsavedIndicator.visible").count() == 1)
    check("en: edited name in dashboard",
          page.locator("#dashboardView .asset-item").first.text_content().__contains__("Test Chequing Account") or
          page.locator("#dashboardView").text_content().__contains__("Test Chequing Account"))

    check("en: with-value count updates after FMV edit", int(page.text_content("#withValueCount")) >= 1)
    page.click('.scope-filter-btn[data-asset-scope="value"]')
    page.wait_for_timeout(200)
    check("en: With Value quick filter narrows to valued assets",
          page.locator("#dashboardView .asset-item").count() == page.evaluate("assets.filter(hasFinancialValue).length"))
    page.click('.scope-filter-btn[data-asset-scope="all"]')
    page.wait_for_timeout(150)

    # Undo / redo via keyboard shortcuts
    page.keyboard.press("Control+z")
    page.wait_for_timeout(200)
    check("en: undo reverts edit",
          "Test Chequing Account" not in page.locator("#dashboardView").text_content())
    page.keyboard.press("Control+Shift+z")
    page.wait_for_timeout(200)
    check("en: redo re-applies edit",
          "Test Chequing Account" in page.locator("#dashboardView").text_content())

    # Charts now show the edited FMV
    _layout(page, "charts")
    check("en: charts populate after FMV edit",
          page.locator("#chartsView svg").count() >= 1)
    page.screenshot(path=str(SHOTS / "en-charts.png"))
    _layout(page, "dashboard")

    # Duplicate
    _open_modal(page, "A-0001")
    page.click("#modalDuplicate")
    page.wait_for_timeout(300)
    dup_id = page.evaluate("currentAsset ? (currentAsset.id || '') : ''")
    check("en: duplicate opens copy in modal", "copy" in dup_id, f"id={dup_id}")
    _save_modal(page)
    check("en: duplicated asset added (518)",
          page.locator("#dashboardView .asset-item").count() == 518)

    # Delete (confirm dialog)
    _open_modal(page, "A-0001-copy")
    page.once("dialog", lambda d: d.accept())
    page.click("#modalDelete")
    page.wait_for_timeout(300)
    check("en: deleted asset removed (517)",
          page.locator("#dashboardView .asset-item").count() == 517)

    # File lock: encrypt the data block (birth date + family word). Enabling
    # stages encrypted data into the DOM script; native Ctrl+S (or the direct
    # download fallback) persists it. Reopening shows an in-page unlock gate.
    _open_modal(page, "A-0001")
    page.fill('#modalContent [data-field="login_password"]', "s3cret-pw!")
    page.fill('#modalContent [data-field="owner"]', "Chen Test")
    _save_modal(page)
    # This legacy File Lock regression validates the universal staged/download
    # fallback. Direct Save has separate Inventory ID/API assertions below.
    page.evaluate("Object.defineProperty(window, 'showSaveFilePicker', {value: undefined, configurable: true})")
    _menu_click(page, "#lockToggle")
    page.wait_for_selector("#lockOverlay.active", timeout=5000)
    status = page.text_content("#lockStatusText")
    check("en: lock not-set status", "plain text" in status.lower() or "明文" in status, status)
    page.fill("#lockBirth", "19750108")
    page.fill("#lockWord1", "chen")
    page.fill("#lockWord2", "chen")
    page.click("#lockEnableBtn")
    page.wait_for_timeout(1200)
    check("en: lock session enabled", page.evaluate("!!lockPassphrase"))
    check("en: save guide shown after lock",
          page.locator("#saveGuideOverlay.active").count() == 1)
    # Direct-download fallback persists the encrypted data block.
    with page.expect_download() as dl:
        page.click("#saveGuideDirect")
    locked_path = Path(dl.value.path())
    locked_html = locked_path.read_text(encoding="utf-8") if locked_path else ""
    check("en: locked download carries encrypted data block",
          '"enc":true' in locked_html, f"size={len(locked_html)}")
    check("en: locked download hides plaintext data",
          "s3cret-pw!" not in locked_html)

    # Reopen the saved (encrypted) file: in-page unlock gate appears.
    lock_tmp = Path(tempfile.mkdtemp()) / "e2e-locked.html"
    lock_tmp.write_text(locked_html, encoding="utf-8")
    lock_ctx = browser.new_context()
    lock_page = lock_ctx.new_page()
    lock_err = []
    lock_page.on("pageerror", lambda e: lock_err.append(str(e)))
    lock_page.goto(lock_tmp.as_uri(), wait_until="load")
    lock_page.wait_for_timeout(500)
    check("en: locked file shows unlock gate",
          lock_page.locator("#lockOverlay.active").count() == 1)
    # Esc/click-away must not dismiss the gate while data is still locked (M4).
    lock_page.keyboard.press("Escape")
    lock_page.wait_for_timeout(200)
    check("en: Escape does not dismiss locked gate",
          lock_page.locator("#lockOverlay.active").count() == 1)
    # Import must be blocked while locked (M4) — otherwise it would write
    # plaintext and silently strip the encryption.
    import_fixture = Path(tempfile.mkdtemp()) / "import.json"
    import_fixture.write_text(
        json.dumps([{"id": "X-1", "asset_name": "Inject", "fmv": 1}]), encoding="utf-8")
    lock_page.set_input_files("#importJSONFile", str(import_fixture))
    lock_page.wait_for_timeout(400)
    check("en: import blocked while locked keeps ciphertext",
          lock_page.evaluate(
              "Array.isArray(INVENTORY_DATA.assets) === false && dataLocked === true"))
    lock_page.fill("#lockUnlockBirth", "19991231")
    lock_page.fill("#lockUnlockWord", "wrong")
    lock_page.click("#lockUnlockBtn")
    lock_page.wait_for_timeout(1200)
    check("en: wrong passphrase rejected (still gated)",
          lock_page.locator("#lockOverlay.active").count() == 1)
    lock_page.fill("#lockUnlockBirth", "19750108")
    lock_page.fill("#lockUnlockWord", "chen")
    lock_page.click("#lockUnlockBtn")
    lock_page.wait_for_timeout(1500)
    check("en: correct passphrase unlocks dashboard",
          lock_page.text_content("#totalAssets") == "517", lock_page.text_content("#totalAssets"))
    check("en: unlocked locked-file carries credentials",
          lock_page.evaluate("assets.find(a => a.id === 'A-0001').login_password") == "s3cret-pw!")
    check("en: owner filter rebuilt after unlock (M3)",
          lock_page.locator("#ownerFilter option").count() > 1
          and lock_page.evaluate(
              "[...document.querySelectorAll('#ownerFilter option')].some(o => o.value === 'Chen Test')"))
    check("en: wizard catalog rebuilt after unlock (M2)",
          lock_page.evaluate("(typeof ASSETS_DATA === 'object') && Array.isArray(ASSETS_DATA) && ASSETS_DATA.length > 0"))
    check("en: locked unlock has no JS errors", not lock_err, f"{lock_err[:3]}")
    lock_ctx.close()
    # Reset the main session to plaintext so the remaining export/save tests
    # exercise the normal (unlocked) path.
    page.evaluate("lockPassphrase = null; dataLocked = false;")

    # Exports (downloads)
    with page.expect_download() as dl:
        _menu_click(page, "#exportMD")
    md_dl = dl.value
    md_content = Path(md_dl.path()).read_text(encoding="utf-8") if md_dl.path() else ""
    check("en: export Markdown downloads", "# Asset Inventory" in md_content)
    check("en: exported MD carries edits", "Test Chequing Account" in md_content)

    with page.expect_download() as dl:
        _menu_click(page, "#exportExcel")
    csv_dl = dl.value
    csv_content = Path(csv_dl.path()).read_text(encoding="utf-8") if csv_dl.path() else ""
    check("en: export CSV downloads", "ID,Category" in csv_content)

    with page.expect_download() as dl:
        _menu_click(page, "#exportJSON")
    json_dl = dl.value
    json_content = Path(json_dl.path()).read_text(encoding="utf-8") if json_dl.path() else ""
    try:
        exported = json.loads(json_content)
        json_ok = isinstance(exported, list) and len(exported) == 517 and \
            any(a.get("asset_name") == "Test Chequing Account" for a in exported)
    except Exception as e:
        json_ok = False
        json_content = str(e)
    check("en: export JSON round-trips 517 assets with edits", json_ok)

    # Save button stages data + shows Ctrl+S guide; direct-download fallback
    # persists a fresh copy with the latest edits.
    page.click("#saveHTML")
    page.wait_for_selector("#saveGuideOverlay.active", timeout=5000)
    check("en: save shows staged/Ctrl+S guide",
          page.locator("#saveGuideOverlay.active").count() == 1)
    with page.expect_download() as dl:
        page.click("#saveGuideDirect")
    html_dl = dl.value
    html_path = html_dl.path()
    saved_html = Path(html_path).read_text(encoding="utf-8") if html_path else ""
    check("en: save-HTML carries edits",
          "const INVENTORY_DATA = {" in saved_html and "Test Chequing Account" in saved_html)
    page.click("#saveGuideClose")
    page.wait_for_timeout(200)

    # Import JSON (replaces assets)
    page.set_input_files("#importJSONFile", str(json_dl.path()))
    page.wait_for_timeout(500)
    check("en: import JSON restores asset set",
          page.locator("#dashboardView .asset-item").count() == 517)

    # Print
    _menu_click(page, "#printBtn")
    page.wait_for_timeout(400)
    print_html = page.text_content("#printView")
    check("en: print view populated", print_html is not None and len(print_html) > 500)
    check("en: emergency access guide is first print section",
          page.locator("#printView").inner_text().find("Emergency Access Guide") >= 0
          and page.locator("#printView").inner_text().find("Emergency Access Guide") < page.locator("#printView").inner_text().find("Master Asset Index"))
    check("en: master asset index printed", "Master Asset Index" in page.locator("#printView").inner_text())
    check("en: print includes inventory id", "INV-" in page.locator("#printView").inner_text())
    page.emulate_media(media="print")
    check("en: print media has no standalone workspace chrome", page.locator(".workspace-bar").count() == 0)
    check("en: print media hides feature menu", page.evaluate("getComputedStyle(document.querySelector('#featureMenu')).display") == "none")
    check("en: print media shows binder", page.locator("#printView").is_visible())
    page.emulate_media(media="screen")

    # ---- Continuity / access readiness ----
    current_inv_id = page.evaluate("INVENTORY_DATA.inventory_id")
    check("en: inventory id format", bool(re.match(r"^INV-[A-F0-9]{12}$", current_inv_id or "")), current_inv_id)
    check("en: distributed template adopts per-family inventory id",
          not str(current_inv_id).startswith("INV-TEMPLATE-"))
    check("en: inventory binding extracts same id",
          page.evaluate("extractInventoryId(document.documentElement.outerHTML)") == current_inv_id)
    check("en: save-as control present", page.locator("#saveAsHTML").count() == 1)
    check("en: persistent binding helpers available",
          page.evaluate("typeof persistSaveFileHandle === 'function' && typeof loadPersistedSaveFileHandle === 'function' && typeof restoreSaveFileHandle === 'function' && typeof ensureWritePermission === 'function'"))
    ready_score = page.evaluate("accessReadiness({institution:'TD',access_location:'vault',access_recovery_contact:'Jane',handoff_instructions:'call TD',incapacity_access:'Ready',death_access:'Ready',last_access_test:new Date().toISOString().slice(0,10)}).score")
    check("en: fully prepared access path scores 100", ready_score == 100, str(ready_score))
    weak_score = page.evaluate("accessReadiness({}).score")
    check("en: empty access path scores critically low", weak_score < 40, str(weak_score))
    _layout(page, "review")
    page.wait_for_timeout(300)
    check("en: annual review renders", page.locator("#reviewView .review-hero").count() == 1)
    check("en: annual review lists readiness rows", page.locator("#reviewView .review-row").count() > 0)
    _layout(page, "dashboard")
    _open_modal(page, "A-0001")
    check("en: handoff fields available in editor",
          all(page.locator(f'#modalContent [data-field="{f}"]').count() == 1 for f in ["emergency_priority", "incapacity_access", "death_access", "handoff_instructions", "last_access_test"]))
    page.click("#modalClose")
    page.wait_for_timeout(200)

    # ---- Quick-add wizard (3.7) ----
    check("en: header Add Asset removed", page.locator("#addAssetBtn").count() == 0)
    page.click("#statAddAssetBtn")
    page.wait_for_selector("#wizardOverlay.active", timeout=5000)
    types = page.locator(".wizard-type-item").count()
    check("en: wizard type list renders", types > 400, f"n={types}")
    page.fill("#wizardSearch", "Chequing")
    page.wait_for_timeout(250)
    narrowed = page.locator(".wizard-type-item").count()
    check("en: wizard search narrows types", 0 < narrowed < types, f"n={narrowed}")
    page.click('.wizard-type-item[data-type="Chequing Account"]')
    page.click("#wizardNext")
    page.wait_for_timeout(250)
    check("en: wizard step 2 institution pre-filled",
          page.locator("#wizardInstitution").input_value() != "")
    page.click("#wizardNext")
    page.wait_for_timeout(250)
    page.fill("#wizardOwner", "James Chen")
    page.click("#wizardNext")
    page.wait_for_timeout(250)
    page.fill("#wizardFmv", "5000")
    page.click("#wizardFinish")
    page.wait_for_selector("#wizardOverlay:not(.active)", timeout=8000)
    check("en: wizard creates asset",
          page.evaluate("assets.some(a => a.source === 'quick-add')"))
    check("en: Assets card increments after quick-add", int(page.text_content("#assetCount")) >= 1, page.text_content("#assetCount"))

    # ---- Bulk edit (3.8) ----
    _layout(page, "table")
    # Earlier sort tests left the table sorted by asset_name desc; restore
    # id-ascending order so A-0001/A-0002 are on page 1 for selection.
    page.click('th[data-sort="id"]')
    page.wait_for_timeout(250)
    page.check('.row-check[data-check-id="A-0001"]')
    page.check('.row-check[data-check-id="A-0002"]')
    check("en: bulk selection bar visible", page.locator("#bulkBar").is_visible())
    page.click("#bulkEditBtn")
    page.wait_for_selector("#bulkOverlay.active", timeout=5000)
    page.select_option("#bulkFieldSelect", "owner")
    page.fill("#bulkValueInput", "Bulk Owner")
    page.click("#bulkApplyBtn")
    page.wait_for_selector("#bulkOverlay:not(.active)", timeout=8000)
    check("en: bulk edit updates both assets",
          page.evaluate("['A-0001', 'A-0002'].every(id => (assets.find(a => a.id === id) || {}).owner === 'Bulk Owner')"))

    # ---- Inline table editing (3.9) ----
    cell = page.locator('#tableBody tr[data-asset-id="A-0001"] td[data-edit-field="asset_name"]')
    cell.dblclick()
    page.wait_for_timeout(250)
    page.fill('#tableBody tr[data-asset-id="A-0001"] .inline-edit-input', "Inline Edited")
    page.keyboard.press("Enter")
    page.wait_for_timeout(300)
    check("en: inline cell edit saved",
          page.evaluate("(assets.find(a => a.id === 'A-0001') || {}).asset_name") == "Inline Edited")

    # ---- Kanban drag-and-drop (3.10) ----
    _layout(page, "kanban")
    check("en: kanban cards draggable",
          page.locator('.kanban-card[data-asset-id="A-0001"]').get_attribute("draggable") == "true")
    page.evaluate("""() => {
        const src = document.querySelector('.kanban-card[data-asset-id="A-0001"]');
        const dst = document.querySelector('.kanban-column[data-kanban-status="Dormant"] .kanban-body');
        const dt = new DataTransfer();
        src.dispatchEvent(new DragEvent('dragstart', { bubbles: true, dataTransfer: dt }));
        dst.dispatchEvent(new DragEvent('dragover', { bubbles: true, cancelable: true, dataTransfer: dt }));
        dst.dispatchEvent(new DragEvent('drop', { bubbles: true, cancelable: true, dataTransfer: dt }));
        src.dispatchEvent(new DragEvent('dragend', { bubbles: true, dataTransfer: dt }));
    }""")
    page.wait_for_timeout(300)
    check("en: kanban drag moves asset to Dormant",
          page.evaluate("(assets.find(a => a.id === 'A-0001') || {}).status") == "Dormant")

    # ---- Column configuration (3.11) ----
    _layout(page, "table")
    page.click("#columnConfigBtn")
    page.wait_for_timeout(250)
    check("en: column menu renders 8 toggles",
          page.locator("#columnConfigMenu label").count() == 8)
    page.uncheck('input[data-col-check="owner"]')
    page.wait_for_timeout(250)
    check("en: owner column hidden",
          page.locator('.data-table th[data-col="owner"]').evaluate("el => el.style.display === 'none'"))
    page.check('input[data-col-check="owner"]')
    page.wait_for_timeout(250)
    check("en: owner column restored",
          page.locator('.data-table th[data-col="owner"]').evaluate("el => el.style.display !== 'none'"))
    page.click("#searchInput")  # close the menu (document click handler)

    # ---- Markdown notes (3.12) ----
    _layout(page, "dashboard")
    _open_modal(page, "A-0001")
    check("en: markdown hint shown for notes field",
          page.locator("#modalContent .markdown-hint").count() >= 3)
    page.fill('#modalContent [data-field="notes"]',
              "**Bold** and *italic* and [link](https://example.com)\n- item1\n- item2")
    _save_modal(page)
    _layout(page, "detail")
    page.click('.detail-asset-item[data-detail-id="A-0001"]')
    page.wait_for_timeout(300)
    detail_html = page.locator("#detailContent").inner_html()
    check("en: markdown bold rendered", "<strong>Bold</strong>" in detail_html)
    check("en: markdown link rendered", '<a href="https://example.com"' in detail_html)

    # ---- Auto-save (extends unsaved-changes indicator) ----
    _layout(page, "dashboard")
    _open_modal(page, "A-0002")
    page.fill('#modalContent [data-field="asset_name"]', "AutoSaved Name")
    _save_modal(page)
    check("en: autosave: dirty dot appears after edit",
          page.locator("#unsavedIndicator.visible").count() == 1)
    page.wait_for_timeout(2500)  # 1.5s debounce + margin
    check("en: autosave: dirty dot clears automatically",
          page.locator("#unsavedIndicator.visible").count() == 0)
    _menu_click(page, "#autoSaveToggle")  # toggle off
    page.wait_for_timeout(250)
    check("en: autosave: toggle-off persisted",
          page.evaluate("localStorage.getItem('autoSaveEnabled')") == "0")
    _open_modal(page, "A-0003")
    page.fill('#modalContent [data-field="asset_name"]', "NoAutosave Name")
    _save_modal(page)
    page.wait_for_timeout(2500)
    check("en: autosave: off keeps dirty dot",
          page.locator("#unsavedIndicator.visible").count() == 1)
    _menu_click(page, "#autoSaveToggle")  # toggle back on
    page.wait_for_timeout(2000)
    check("en: autosave: on clears dot",
          page.locator("#unsavedIndicator.visible").count() == 0)

    # ---- Regression checks (code-review fixes) ----
    # CSV formula-injection guard rejects leading-whitespace formulas
    guard_ok = page.evaluate("""(() => {
        const bad = [' =cmd|xyz', '+SUM(1,2)', '@link', '-123'];
        return bad.every(v => !/^[=+\-@]/.test(csvEscape(v).replace(/^"/, ''))) &&
               !/^[=+\-@]/.test(csvEscape('normal'));
    })()""")
    check("en: csvEscape guards leading-space formula", guard_ok,
          page.evaluate("csvEscape(' =cmd|xyz')"))

    # Audit view follows the active search/filter
    _layout(page, "audit")
    n_all_audit = page.locator(".audit-item").count()
    page.fill("#searchInput", "Chequing")
    page.wait_for_timeout(600)
    n_filt_audit = page.locator(".audit-item").count()
    check("en: audit follows search filter", 0 < n_filt_audit < n_all_audit,
          f"n={n_filt_audit}/{n_all_audit}")
    page.fill("#searchInput", "")
    page.wait_for_timeout(600)

    # Compact category collapse hides its assets
    _layout(page, "compact")
    page.locator(".compact-category-header").first.click()
    page.wait_for_timeout(250)
    check("en: compact category collapse toggles body",
          page.locator(".compact-category-body.collapsed").count() >= 1)
    page.locator(".compact-category-header").first.click()  # restore
    page.wait_for_timeout(250)

    # Numeric-id JSON import is normalized to strings and rendered safely
    with tempfile.TemporaryDirectory() as tmp2:
        num_assets = json.loads(json_content)
        for i, a in enumerate(num_assets):
            a["id"] = i + 1  # numeric ids, as some exports/imports may produce
        num_path = Path(tmp2) / "numeric-ids.json"
        num_path.write_text(json.dumps(num_assets), encoding="utf-8")
        page.set_input_files("#importJSONFile", str(num_path))
        page.wait_for_timeout(500)
        ok_types = page.evaluate("assets.every(a => typeof a.id === 'string')")
        check("en: numeric-id import normalized to strings", ok_types,
              page.evaluate("assets.slice(0,3).map(a => a.id).join(',')"))
        _layout(page, "dashboard")
        check("en: numeric-id import renders without crash",
              page.locator("#dashboardView .asset-item").count() == len(num_assets))
        # duplicate ids are rejected, keeping the previous set
        errs_before_dup = list(js_errors)  # the reject logs an expected console.error
        dup_assets = json.loads(json_content)
        dup_assets[1]["id"] = dup_assets[0]["id"]
        dup_path = Path(tmp2) / "duplicate-ids.json"
        dup_path.write_text(json.dumps(dup_assets), encoding="utf-8")
        page.set_input_files("#importJSONFile", str(dup_path))
        page.wait_for_timeout(500)
        check("en: duplicate-id import rejected (state preserved)",
              page.evaluate("assets.length") == len(num_assets),
              page.evaluate("String(assets.length)"))
        # drop the expected duplicate-rejection console.error from the list
        js_errors[:] = errs_before_dup
        restore_path = Path(tmp2) / "restore.json"
        restore_path.write_text(json_content, encoding="utf-8")
        page.set_input_files("#importJSONFile", str(restore_path))
        page.wait_for_timeout(500)
        check("en: original export re-imported to restore state",
              page.locator("#dashboardView .asset-item").count() == 517)

    # No JS errors across the whole session
    check("en: zero JS errors on page", not js_errors, f"{js_errors[:5]}")

    context.close()
    return ok


def e2e_zh(browser):
    print("  -- zh dashboard --")
    context, page, js_errors = _new_page(browser, ZH_HTML)
    ok = True

    check("zh: title is Chinese", "资产清单" in page.title(), page.title())
    check("zh: stat totalAssets = 517", page.text_content("#totalAssets") == "517")
    check("zh: search placeholder translated",
          "搜索" in page.get_attribute("#searchInput", "placeholder"))
    header_labels = page.locator("#headerViewSwitcher .header-layout-btn").evaluate_all(
        "els => els.map(el => el.getAttribute('aria-label'))")
    check("zh: Professional header views use bilingual terminology",
          "Audit 财产审计" in header_labels and "Annual Review 年度复核" in header_labels)
    _open_feature_menu(page)
    menu_text = page.locator("#featureMenu").text_content()
    check("zh: Professional export section localized", "Export 导出" in menu_text)
    page.click("#featureMenuClose")
    check("zh: dashboard view renders 517 items",
          page.locator("#dashboardView .asset-item").count() == 517)
    check("zh: no leftover TR_ placeholders in body",
          "{{TR_" not in page.locator("body").inner_text())
    # Interact: search + theme + a layout. asset_name stays English in zh
    # builds; search by a translated category (加密货币 = Crypto) instead.
    page.fill("#searchInput", "加密货币")
    page.wait_for_timeout(600)
    n = page.locator("#dashboardView .asset-item").count()
    check("zh: Chinese search works", 0 < n < 517, f"n={n}")
    page.fill("#searchInput", "")
    _menu_click(page, "#themeToggle")
    theme = page.evaluate("document.documentElement.getAttribute('data-theme')")
    check("zh: theme toggle works", theme == "dark", theme)
    _layout(page, "audit")
    check("zh: audit renders in Chinese", page.locator(".audit-item").count() == 517)
    page.screenshot(path=str(SHOTS / "zh-dashboard.png"))
    check("zh: zero JS errors on page", not js_errors, f"{js_errors[:5]}")

    context.close()
    return ok


# =============================================================================

def main():
    SHOTS.mkdir(parents=True, exist_ok=True)
    run_generation()
    static_checks()
    demo_fixture_check()
    tier_gating_check()
    license_downgrade_ui_check()
    responsive_ui_check()
    browser_e2e()
    import shutil
    shutil.rmtree(TMP, ignore_errors=True)

    print("\n" + "=" * 60)
    print(f"RESULT: {len(PASSED)} passed, {len(FAILED)} failed")
    if FAILED:
        print("Failures:")
        for f in FAILED:
            print(f"  - {f}")
    print("=" * 60)
    return 0 if not FAILED else 1


if __name__ == "__main__":
    sys.exit(main())
